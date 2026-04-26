import os
import sys
import re
import html
import math
import random
import json
import shutil
import sqlite3
from datetime import datetime
from typing import Any, Dict, List, Optional

from PySide6.QtWidgets import (
     QMainWindow, QSplitter, QStackedWidget, QListWidget, QWidget,
    QVBoxLayout, QHBoxLayout, QGroupBox, QPushButton, QLineEdit, QLabel,
    QTableWidget, QTableWidgetItem, QGraphicsView, QGraphicsScene,
    QGraphicsEllipseItem,  QGraphicsLineItem,QTextEdit,
    QMessageBox, QMenu,  QFileDialog, QDialog, QFormLayout,
    QComboBox, QSpinBox, QHeaderView, QInputDialog, QDateTimeEdit, QFrame,
    QScrollArea,
)
from PySide6.QtCore import Qt, QTimer, QPointF, QDateTime, QRect
from PySide6.QtGui import (
    QColor,
    QPen,
    QBrush,
    QPainter,
    QFont,
    QPolygonF,
    QPixmap,
    QPainterPath,
    QLinearGradient,
    QIcon,
    QImageReader,
    QGuiApplication,
)

from database import Database, InventoryManager
from auth_service import hash_password
from warehouse_manager import WarehouseManager
from material_manager import MaterialManager
from dispatch_manager import DispatchManager


class WarehouseDispatchRecordsDialog(QDialog):
    """仓库节点右键「调度记录」：大窗口 + 每条调度单独卡片分隔。"""

    def __init__(
        self,
        warehouse_name: str,
        warehouse_id: int,
        records: List[Dict[str, Any]],
        bulk: Dict[int, List[Dict[str, Any]]],
        parent: Optional[QWidget] = None,
    ):
        super().__init__(parent)
        self.setWindowTitle(f"{warehouse_name} — 调度记录")
        self.resize(820, 560)
        self.setMinimumSize(640, 400)
        self.setStyleSheet("QDialog { background-color: #f0f2f5; }")

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        tip = QLabel(f"共 {len(records)} 条调度（可滚动查看）")
        tip.setStyleSheet("font-size: 13px; color: #606266;")
        root.addWidget(tip)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; }")

        inner = QWidget()
        inner.setStyleSheet("background: transparent;")
        v = QVBoxLayout(inner)
        v.setSpacing(14)
        v.setContentsMargins(4, 4, 8, 8)

        for r in records:
            rid = int(r["id"])
            items = bulk.get(rid, [])
            mat_raw = DispatchManager.format_items_summary(items) if items else ""
            mat_cell = _format_dispatch_materials_qty_display(mat_raw) if mat_raw else "无"
            role = "调出" if r.get("from_warehouse_id") == warehouse_id else "调入"
            head = (
                f"#{rid}　[{html.escape(role)}]　"
                f"{html.escape(str(r.get('from_name') or ''))} → {html.escape(str(r.get('to_name') or ''))}"
            )
            sub = (
                f"{html.escape(str(r.get('timestamp') or ''))}　"
                f"执行人：{html.escape(str(r.get('executor') or ''))}"
            )
            mat_html = html.escape(mat_cell).replace("\n", "<br/>")

            card = QFrame()
            card.setObjectName("dispatchRecordCard")
            card.setStyleSheet(
                """
                QFrame#dispatchRecordCard {
                    background-color: #ffffff;
                    border: 1px solid #dcdfe6;
                    border-radius: 10px;
                }
                """
            )
            cv = QVBoxLayout(card)
            cv.setContentsMargins(14, 12, 14, 12)
            cv.setSpacing(6)
            hl = QLabel(head)
            hl.setTextFormat(Qt.RichText)
            hl.setStyleSheet("font-size: 14px; font-weight: bold; color: #303133;")
            hl.setWordWrap(True)
            cv.addWidget(hl)
            sl = QLabel(sub)
            sl.setStyleSheet("font-size: 12px; color: #909399;")
            sl.setWordWrap(True)
            cv.addWidget(sl)
            lab_m = QLabel("物料：")
            lab_m.setStyleSheet(
                "font-size: 13px; font-weight: bold; color: #606266; margin-top: 4px;"
            )
            cv.addWidget(lab_m)
            ml = QLabel()
            ml.setTextFormat(Qt.RichText)
            ml.setText(mat_html)
            ml.setStyleSheet("font-size: 13px; color: #409eff; line-height: 1.45; padding-left: 2px;")
            ml.setWordWrap(True)
            cv.addWidget(ml)
            v.addWidget(card)

        v.addStretch()
        scroll.setWidget(inner)
        root.addWidget(scroll, stretch=1)

        row_btn = QHBoxLayout()
        row_btn.addStretch()
        btn = QPushButton("关闭")
        btn.setMinimumWidth(120)
        btn.setStyleSheet(
            "QPushButton { background: #409eff; color: white; border: none; "
            "border-radius: 6px; padding: 10px 24px; font-size: 14px; }"
            "QPushButton:hover { background: #66b1ff; }"
        )
        btn.clicked.connect(self.accept)
        row_btn.addWidget(btn)
        root.addLayout(row_btn)


def _dev_source_root() -> str:
    """含 main.py 与 ui 包的项目根（开发与源码运行）。"""
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def bundle_root() -> str:
    """只读资源：PyInstaller 展开目录（_internal）或开发时项目根。"""
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass and os.path.isdir(meipass):
            return meipass
        exe_dir = os.path.dirname(sys.executable)
        internal = os.path.join(exe_dir, "_internal")
        if os.path.isdir(internal):
            return internal
        return exe_dir
    return _dev_source_root()


def user_data_root() -> str:
    """可写数据：打包后与 exe 同目录；开发时与项目根相同。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return _dev_source_root()


def asset_path(*parts: str) -> str:
    """静态资源：优先 exe 旁覆盖，否则使用打包进 _MEIPASS 的副本。"""
    rel = os.path.join(*parts)
    user = os.path.join(user_data_root(), rel)
    if os.path.isfile(user) or os.path.isdir(user):
        return user
    return os.path.join(bundle_root(), rel)


def window_icon_path() -> Optional[str]:
    """主窗口 / 登录框图标（assets/app.ico），不存在则返回 None。"""
    p = asset_path("assets", "app.ico")
    return p if os.path.isfile(p) else None


def window_icon_qicon() -> QIcon:
    """从 app.ico 构建多尺寸 QIcon（任务栏/标题栏在高 DPI 下更稳定）；失败则回退为单路径加载。"""
    p = window_icon_path()
    if not p:
        return QIcon()
    reader = QImageReader(p)
    n = reader.imageCount()
    if n > 0:
        icon = QIcon()
        for i in range(n):
            if not reader.jumpToImage(i):
                break
            img = reader.read()
            if img.isNull():
                continue
            icon.addPixmap(QPixmap.fromImage(img))
        if not icon.isNull():
            return icon
    return QIcon(p)


def bootstrap_frozen_resources() -> None:
    """仅 wellsite.db 与 Picture record 使用共享路径；布局与备份页配置在本机 exe 旁，无需从包内复制。"""
    return


NETWORK_SETTINGS_FILE = os.path.join(user_data_root(), "network_settings.json")


def _read_network_settings_file() -> dict:
    if not os.path.isfile(NETWORK_SETTINGS_FILE):
        return {}
    try:
        with open(NETWORK_SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def network_db_path_from_settings() -> str:
    from runtime_flags import resolved_shared_database_path

    return resolved_shared_database_path()


def is_network_data_enabled() -> bool:
    from runtime_flags import resolved_shared_database_path

    return bool(resolved_shared_database_path())


def shared_data_root() -> str:
    """与 wellsite.db 同目录（共享 UNC 根）。仅用于 Picture record 等与库同机路径。"""
    from runtime_flags import shared_pack_data_root

    return shared_pack_data_root()


def backup_ui_settings_file() -> str:
    return os.path.join(user_data_root(), "backup_ui_settings.json")


def dispatch_photo_dir() -> str:
    r = shared_data_root()
    return os.path.join(r, "Picture record") if r else ""


def warehouse_layout_file() -> str:
    return os.path.join(user_data_root(), "warehouse_layout.json")


def shared_background_dir() -> str:
    return os.path.join(user_data_root(), "Background images")


def copy_file_to_shared(src_path: str, subdir: str) -> str:
    """将文件复制到本机数据目录（exe 旁）下的子目录，例如 Background images。"""
    r = user_data_root()
    ap = os.path.abspath(os.path.normpath(src_path))
    root = os.path.abspath(os.path.join(r, subdir))
    os.makedirs(root, exist_ok=True)
    try:
        if os.path.normcase(ap).startswith(os.path.normcase(root + os.sep)):
            return ap
    except (OSError, ValueError):
        pass
    base = os.path.basename(ap)
    safe = re.sub(r"[^\w\-\.()\u4e00-\u9fff]", "_", base).strip("._") or "file"
    stem, ext = os.path.splitext(safe)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(root, f"{ts}_{stem}{ext}")
    n = 0
    while os.path.exists(dest):
        n += 1
        dest = os.path.join(root, f"{ts}_{stem}_{n}{ext}")
    shutil.copy2(ap, dest)
    return os.path.abspath(dest)


def store_dispatch_photo(src_path: Optional[str]) -> Optional[str]:
    """
    将调度附图复制到与 wellsite.db 同目录下的 Picture record，返回绝对路径写入数据库。
    若已在该目录内则直接返回原路径；无图、源不存在或未配置共享库路径则返回 None。
    """
    if not src_path:
        return None
    ap = os.path.abspath(os.path.normpath(src_path))
    if not os.path.isfile(ap):
        return None
    dest_root = dispatch_photo_dir()
    if not dest_root:
        return None
    dest_root = os.path.abspath(dest_root)
    os.makedirs(dest_root, exist_ok=True)
    try:
        if os.path.normcase(ap).startswith(os.path.normcase(dest_root + os.sep)):
            return ap
    except (OSError, ValueError):
        pass
    base = os.path.basename(ap)
    safe = re.sub(r"[^\w\-\.()\u4e00-\u9fff]", "_", base).strip("._") or "image"
    stem, ext = os.path.splitext(safe)
    if not ext:
        ext = ".jpg"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(dest_root, f"{ts}_{stem}{ext}")
    n = 0
    while os.path.exists(dest):
        n += 1
        dest = os.path.join(dest_root, f"{ts}_{stem}_{n}{ext}")
    shutil.copy2(ap, dest)
    return os.path.abspath(dest)


# 与备份页背景样式合并使用，避免 _apply_backup_page_background 覆盖后按钮恢复全局 QSS 重影
_BACKUP_PAGE_BTN_QSS = """
    QWidget#backupPage QPushButton#backupOutlineBtn {
        background-color: rgba(64, 158, 255, 0.14);
        color: #409eff;
        border: 1px solid #409eff;
        border-radius: 6px;
        padding: 10px 18px;
        min-height: 22px;
        font-size: 13px;
        font-weight: normal;
    }
    QWidget#backupPage QPushButton#backupOutlineBtn:hover {
        background-color: rgba(64, 158, 255, 0.22);
    }
    QWidget#backupPage QPushButton#backupMainBtn {
        background-color: #409eff;
        color: #ffffff;
        border: none;
        border-radius: 8px;
        padding: 12px 24px;
        min-height: 24px;
        font-size: 15px;
        font-weight: normal;
    }
    QWidget#backupPage QPushButton#backupMainBtn:hover {
        background-color: #66b1ff;
    }
"""


class BackupPageRoot(QWidget):
    """
    数据备份页根控件：用 paintEvent 绘制背景图（比 QSS border-image 更稳，与是否缺少 assets 文件夹无关）。
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("backupPage")
        self._bg_pixmap: Optional[QPixmap] = None

    def reload_background(self, main_window: "MainWindow") -> None:
        self._bg_pixmap = None
        s = main_window._read_backup_ui_settings()
        custom = (s.get("background_image") or "").strip()
        if custom:
            ap = os.path.abspath(os.path.normpath(custom))
            if os.path.isfile(ap):
                pm = QPixmap(ap)
                if not pm.isNull():
                    self._bg_pixmap = pm
        if self._bg_pixmap is None and not is_network_data_enabled():
            for fname in ("backup_bg.jpg", "backup_bg.png", "home_bg.jpg", "home_bg.png"):
                p = asset_path("assets", fname)
                if os.path.isfile(p):
                    pm = QPixmap(p)
                    if not pm.isNull():
                        self._bg_pixmap = pm
                        break
        self.setStyleSheet(_BACKUP_PAGE_BTN_QSS)
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), QColor("#eef2f6"))
        if self._bg_pixmap is not None and not self._bg_pixmap.isNull():
            sc = self._bg_pixmap.scaled(
                self.size(),
                Qt.KeepAspectRatioByExpanding,
                Qt.SmoothTransformation,
            )
            x = max(0, (sc.width() - self.width()) // 2)
            y = max(0, (sc.height() - self.height()) // 2)
            painter.drawPixmap(0, 0, sc, x, y, self.width(), self.height())
        else:
            grad = QLinearGradient(0.0, 0.0, float(self.width()), float(self.height()))
            grad.setColorAt(0.0, QColor("#1e3c72"))
            grad.setColorAt(0.5, QColor("#2a5298"))
            grad.setColorAt(1.0, QColor("#7e8ba3"))
            painter.fillRect(self.rect(), QBrush(grad))
        super().paintEvent(event)


def _format_dispatch_materials_qty_display(text: str) -> str:
    """将汇总串里的「 x5.0」改为「 x5」；非整数仍显示最简小数。"""
    if not text:
        return text

    def repl(match) -> str:
        raw = match.group(1)
        try:
            v = float(raw)
            if v == int(v):
                return " x" + str(int(v))
            s = str(v).rstrip("0").rstrip(".")
            return " x" + (s if s else "0")
        except ValueError:
            return match.group(0)

    return re.sub(r" x([\d.]+)", repl, text)


class FlowEdge(QGraphicsLineItem):
    def __init__(self, source: "WarehouseNode", target: "WarehouseNode"):
        super().__init__()
        self.source = source
        self.target = target

        # 线条样式
        self.setPen(QPen(QColor("#00ffff"), 2))

        # 动画点
        self.flow_pos = 0

        self.timer = QTimer()
        self.timer.timeout.connect(self.animate)
        self.timer.start(50)

    def update_position(self):
        """更新线位置"""
        p1 = self.source.scenePos() + QPointF(self.source.size / 2, self.source.size / 2)
        p2 = self.target.scenePos() + QPointF(self.target.size / 2, self.target.size / 2)
        self.setLine(p1.x(), p1.y(), p2.x(), p2.y())

    def animate(self):
        """流动动画"""
        self.flow_pos += 0.02
        if self.flow_pos > 1:
            self.flow_pos = 0
        self.update()

    def paint(self, painter, option, widget):
        self.update_position()

        super().paint(painter, option, widget)

        # 画流动点
        line = self.line()
        x = line.x1() + (line.x2() - line.x1()) * self.flow_pos
        y = line.y1() + (line.y2() - line.y1()) * self.flow_pos

        painter.setBrush(QBrush(QColor("#00ffcc")))
        painter.drawEllipse(QPointF(x, y), 5, 5)

        # 画箭头
        angle = math.atan2(line.dy(), line.dx())
        arrow_size = 10

        p2 = QPointF(line.x2(), line.y2())

        p1 = p2 - QPointF(math.cos(angle - 0.3) * arrow_size,
                          math.sin(angle - 0.3) * arrow_size)
        p3 = p2 - QPointF(math.cos(angle + 0.3) * arrow_size,
                          math.sin(angle + 0.3) * arrow_size)

        painter.setBrush(QColor("#00ffff"))
        painter.drawPolygon(p2, p1, p3)

# ------------------ 仓库节点 ------------------
class WarehouseNode(QGraphicsEllipseItem):
    def __init__(self, warehouse_id, name, x, y, color="#00c8ff", size=120,
                 inventory=0, dispatch_records=None, shape="circle", font_size=10, avatar_path=""):
        super().__init__(0, 0, size, size)
        self.warehouse_id = warehouse_id
        self.name = name
        self.setPos(x, y)
        self.saved_pos = QPointF(x, y)
        self.size = size
        self.node_shape = shape
        self.font_size = font_size
        self.avatar_path = avatar_path or ""
        self.avatar_pixmap = QPixmap(self.avatar_path) if self.avatar_path and os.path.exists(self.avatar_path) else QPixmap()
        self.inventory = inventory
        self.dispatch_records = dispatch_records or []

        self.setBrush(QBrush(QColor(color)))
        self.setPen(QColor("#0ff"))
        self.setFlag(QGraphicsEllipseItem.ItemIsMovable, True)
        self.setFlag(QGraphicsEllipseItem.ItemIsSelectable, True)
        self.setAcceptedMouseButtons(Qt.LeftButton | Qt.RightButton)

    # 拖拽结束保存位置
    def mouseReleaseEvent(self, event):
        super().mouseReleaseEvent(event)
        self.saved_pos = self.pos()
        # 保存到磁盘
        if hasattr(self.scene(), "save_layout_to_disk"):
            self.scene().save_layout_to_disk()

    # 重写绘制方法
    def paint(self, painter: QPainter, option, widget=None):
        painter.setBrush(Qt.NoBrush)
        painter.setPen(self.pen())
        rect = self.rect()
        shape_path = QPainterPath()
        if self.node_shape == "square":
            shape_path.addRect(rect)
        elif self.node_shape == "rounded":
            shape_path.addRoundedRect(rect, 16, 16)
        elif self.node_shape == "diamond":
            cx = rect.x() + rect.width() / 2
            cy = rect.y() + rect.height() / 2
            points = QPolygonF([
                QPointF(cx, rect.y()),
                QPointF(rect.x() + rect.width(), cy),
                QPointF(cx, rect.y() + rect.height()),
                QPointF(rect.x(), cy),
            ])
            shape_path.addPolygon(points)
        else:
            shape_path.addEllipse(rect)

        # 先画头像（如果有），再画轮廓线
        painter.save()
        painter.setClipPath(shape_path)
        if not self.avatar_pixmap.isNull():
            painter.drawPixmap(rect.toRect(), self.avatar_pixmap)
        else:
            painter.fillPath(shape_path, self.brush())
        painter.restore()
        painter.drawPath(shape_path)

        # 绘制文字
        painter.setPen(Qt.white)
        font = QFont("Microsoft YaHei", int(self.font_size), QFont.Bold)
        painter.setFont(font)
        metrics = painter.fontMetrics()
        text_width = metrics.horizontalAdvance(self.name)
        text_height = metrics.height()

        painter.drawText(
            rect.x() + (rect.width() - text_width) / 2,
            rect.y() + (rect.height() + text_height) / 2 - metrics.descent(),
            self.name
        )

    # 右键菜单
    def contextMenuEvent(self, event):
        menu = QMenu()
        menu.setFont(QFont("Microsoft YaHei", 12))
        menu.setMinimumWidth(180)
        menu.setStyleSheet(
            """
            QMenu {
                padding: 2px 0px;
            }
            QMenu::item {
                padding: 4px 18px 4px 10px;
                min-height: 20px;
            }
            QMenu::item:selected {
                background-color: #ecf5ff;
                color: #409eff;
            }
            """
        )
        change_color_action = menu.addAction("🔹 修改颜色")
        upload_avatar_action = menu.addAction("🖼 上传头像")
        clear_avatar_action = menu.addAction("🧹 清除头像")
        change_shape_action = menu.addAction("🔷 修改形状")
        change_size_action = menu.addAction("📏 修改大小")
        change_font_action = menu.addAction("🔤 修改字体大小")
        view_inventory_action = menu.addAction("📦 查看库存")
        view_dispatch_action = menu.addAction("🚛 调度记录")
        action = menu.exec(event.screenPos())
        if action == change_color_action:
            self.change_color()
        elif action == upload_avatar_action:
            self.upload_avatar()
        elif action == clear_avatar_action:
            self.clear_avatar()
        elif action == change_shape_action:
            self.change_shape()
        elif action == change_size_action:
            self.change_size()
        elif action == change_font_action:
            self.change_font_size()
        elif action == view_inventory_action:
            self.show_inventory()
        elif action == view_dispatch_action:
            self.show_dispatch_records()

    def change_color(self):
        from PySide6.QtWidgets import QColorDialog
        color = QColorDialog.getColor()
        if color.isValid():
           self.setBrush(QBrush(color))
           # 保存到磁盘
           if hasattr(self.scene(), "save_layout_to_disk"):
               self.scene().save_layout_to_disk()

    def upload_avatar(self):
        path, _ = QFileDialog.getOpenFileName(
            None,
            "选择头像",
            "",
            "Image Files (*.png *.jpg *.jpeg *.bmp *.webp)",
        )
        if not path:
            return
        pix = QPixmap(path)
        if pix.isNull():
            QMessageBox.warning(None, "提示", "图片加载失败，请选择有效图片")
            return
        self.avatar_path = path
        self.avatar_pixmap = pix
        self.update()
        if hasattr(self.scene(), "save_layout_to_disk"):
            self.scene().save_layout_to_disk()

    def clear_avatar(self):
        self.avatar_path = ""
        self.avatar_pixmap = QPixmap()
        self.update()
        if hasattr(self.scene(), "save_layout_to_disk"):
            self.scene().save_layout_to_disk()

    def change_shape(self):
        options = ["circle", "square", "rounded", "diamond", "pill"]
        labels = {
            "circle": "圆形",
            "square": "方形",
            "rounded": "圆角矩形",
            "diamond": "菱形",
            "pill": "胶囊形",
        }
        current_label = labels.get(self.node_shape, "圆形")
        ui_options = [labels[o] for o in options]
        selected, ok = QInputDialog.getItem(
            None,
            "修改形状",
            "请选择节点形状：",
            ui_options,
            ui_options.index(current_label) if current_label in ui_options else 0,
            False,
        )
        if not ok:
            return
        reverse_labels = {v: k for k, v in labels.items()}
        self.node_shape = reverse_labels.get(selected, "circle")
        self.update()
        if hasattr(self.scene(), "save_layout_to_disk"):
            self.scene().save_layout_to_disk()

    def change_size(self):
        value, ok = QInputDialog.getInt(
            None,
            "修改大小",
            "请输入节点大小（像素）：",
            int(self.size),
            60,
            300,
            5,
        )
        if not ok:
            return
        self.size = value
        self.setRect(0, 0, self.size, self.size)
        self.update()
        if hasattr(self.scene(), "save_layout_to_disk"):
            self.scene().save_layout_to_disk()

    def change_font_size(self):
        value, ok = QInputDialog.getInt(
            None,
            "修改字体大小",
            "请输入字体大小：",
            int(self.font_size),
            8,
            32,
            1,
        )
        if not ok:
            return
        self.font_size = value
        self.update()
        if hasattr(self.scene(), "save_layout_to_disk"):
            self.scene().save_layout_to_disk()

    def show_inventory(self):
        QMessageBox.information(None, f"{self.name} 库存", f"库存数量: {self.inventory}")

    def show_dispatch_records(self):
        scene = self.scene()
        mgr = getattr(scene, "dispatch_mgr", None) if scene is not None else None
        if mgr is None:
            QMessageBox.warning(
                None,
                f"{self.name} 调度记录",
                "调度数据未就绪，请切换到「调度管理」页面后再试。",
            )
            return
        wid = self.warehouse_id
        try:
            wid = int(wid)
        except (TypeError, ValueError):
            pass
        records = mgr.list_dispatches_for_warehouse(wid)
        if not records:
            QMessageBox.information(None, f"{self.name} 调度记录", "暂无调度记录")
            return
        ids = [int(r["id"]) for r in records]
        bulk = mgr.list_dispatch_items_for_records(ids)
        dlg = WarehouseDispatchRecordsDialog(self.name, wid, records, bulk, None)
        dlg.exec()

# ------------------ 拓扑场景 ------------------
class TopologyScene(QGraphicsScene):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.nodes = {}
        self.saved_layout = {}
        self.background_image = ""
        # 从 warehouse_layout.json 的 __main_window__ 读出，供 MainWindow 启动时恢复
        self.pending_main_window_geometry: Optional[Dict[str, Any]] = None
        # 从 __scene__.dispatch_splitter_sizes 读出：拓扑图与调度记录表格的上下分割比例
        self.pending_dispatch_splitter_sizes: Optional[List[int]] = None
        # 由 MainWindow 在创建调度页后注入，供节点右键「调度记录」查询数据库
        self.dispatch_mgr: Optional[Any] = None

    def build(self, warehouses: list[dict]):
        self.clear()
        self.nodes.clear()
        scene_width = 1200
        scene_height = 600
        self.setSceneRect(0, 0, scene_width, scene_height)

        for wh in warehouses:
            wh_id = str(wh['id'])
            if wh_id in self.saved_layout:
                pos = self.saved_layout[wh_id]
                x, y = pos['x'], pos['y']
                color = pos.get('color', wh.get('color', "#00c8ff"))
                size = pos.get('size', wh.get('size', 120))
                shape = pos.get('shape', wh.get('shape', "circle"))
                font_size = pos.get('font_size', wh.get('font_size', 10))
                avatar_path = pos.get('avatar_path', wh.get('avatar_path', ""))
            else:
                x = random.uniform(0, scene_width - wh.get('size', 120))
                y = random.uniform(0, scene_height - wh.get('size', 120))
                color = wh.get('color', "#00c8ff")
                size = wh.get('size', 120)
                shape = wh.get('shape', "circle")
                font_size = wh.get('font_size', 10)
                avatar_path = wh.get('avatar_path', "")

            node = WarehouseNode(
                wh['id'], wh['name'], x, y, color=color, size=size, shape=shape, font_size=font_size, avatar_path=avatar_path
            )
            self.addItem(node)
            self.nodes[wh['id']] = node
            self.saved_layout[wh_id] = {
                "x": x,
                "y": y,
                "color": color,
                "size": size,
                "shape": shape,
                "font_size": font_size,
                "avatar_path": avatar_path,
            }

    @staticmethod
    def _layout_file_read() -> str:
        user = warehouse_layout_file()
        if user and os.path.isfile(user):
            return user
        bundled = os.path.join(bundle_root(), "warehouse_layout.json")
        if os.path.isfile(bundled):
            return bundled
        return user or bundled

    @staticmethod
    def _layout_file_write() -> str:
        return warehouse_layout_file()

    def save_layout_to_disk(
        self,
        main_window: Optional[Any] = None,
        dispatch_splitter_sizes: Optional[List[int]] = None,
    ):
        """保存所有节点位置和颜色到文件；可选写入主窗口几何与调度页上下分割条高度。"""
        data = {}
        for wh_id, node in self.nodes.items():
            data[wh_id] = {
                "x": node.pos().x(),
                "y": node.pos().y(),
                "color": node.brush().color().name(),
                "size": node.size,
                "shape": node.node_shape,
                "font_size": node.font_size,
                "avatar_path": node.avatar_path,
            }

        mw_state: Optional[Dict[str, Any]] = None
        old_scene: Dict[str, Any] = {}
        read_path = self._layout_file_read()
        if os.path.isfile(read_path):
            try:
                with open(read_path, "r", encoding="utf-8") as fr:
                    old = json.load(fr)
                raw = old.get("__main_window__")
                if isinstance(raw, dict):
                    mw_state = raw
                osc = old.get("__scene__")
                if isinstance(osc, dict):
                    old_scene = osc
            except (json.JSONDecodeError, OSError, TypeError):
                pass

        scene_out: Dict[str, Any] = {"background_image": self.background_image}
        dss_old = old_scene.get("dispatch_splitter_sizes")
        if isinstance(dss_old, list) and len(dss_old) >= 2:
            try:
                scene_out["dispatch_splitter_sizes"] = [int(dss_old[0]), int(dss_old[1])]
            except (TypeError, ValueError):
                pass
        if dispatch_splitter_sizes is not None and len(dispatch_splitter_sizes) >= 2:
            try:
                scene_out["dispatch_splitter_sizes"] = [
                    int(dispatch_splitter_sizes[0]),
                    int(dispatch_splitter_sizes[1]),
                ]
            except (TypeError, ValueError):
                pass
        data["__scene__"] = scene_out

        if main_window is not None:
            try:
                if main_window.isMaximized():
                    g = main_window.normalGeometry()
                else:
                    g = main_window.geometry()
                mw_state = {
                    "x": int(g.x()),
                    "y": int(g.y()),
                    "width": int(g.width()),
                    "height": int(g.height()),
                    "maximized": bool(main_window.isMaximized()),
                }
            except Exception:
                pass

        if mw_state:
            data["__main_window__"] = mw_state

        # 同步内存缓存，避免刷新时使用旧布局导致节点回退
        self.saved_layout = {
            str(k): v for k, v in data.items() if k not in ("__scene__", "__main_window__")
        }
        path = self._layout_file_write()
        if not path:
            return
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def load_layout_from_disk(self):
        """读取布局和颜色，如果文件存在"""
        path = self._layout_file_read()
        self.pending_main_window_geometry = None
        self.pending_dispatch_splitter_sizes = None
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                self.saved_layout = json.load(f)
            scene_cfg = self.saved_layout.pop("__scene__", {})
            self.background_image = scene_cfg.get("background_image", "")
            dss = scene_cfg.get("dispatch_splitter_sizes")
            if isinstance(dss, list) and len(dss) >= 2:
                try:
                    self.pending_dispatch_splitter_sizes = [int(dss[0]), int(dss[1])]
                except (TypeError, ValueError):
                    self.pending_dispatch_splitter_sizes = None
            mw = self.saved_layout.pop("__main_window__", None)
            self.pending_main_window_geometry = mw if isinstance(mw, dict) else None
        else:
            self.saved_layout = {}
            self.background_image = ""

    def set_background_image(self, path: str):
        self.background_image = path or ""
        self.save_layout_to_disk()
        self.update()

    def drawBackground(self, painter: QPainter, rect):
        super().drawBackground(painter, rect)
        if not self.background_image or not os.path.exists(self.background_image):
            return
        pix = QPixmap(self.background_image)
        if pix.isNull():
            return
        # 按当前可视窗口范围自适应铺满背景
        views = self.views()
        if views:
            visible_rect = views[0].mapToScene(views[0].viewport().rect()).boundingRect()
            painter.drawPixmap(visible_rect.toRect(), pix)
        else:
            painter.drawPixmap(self.sceneRect().toRect(), pix)



class MainWindow(QMainWindow):
    def __init__(
        self,
        db: Database,
        warehouse_mgr: WarehouseManager,
        material_mgr: MaterialManager,
        dispatch_mgr: DispatchManager,
        current_user: Dict[str, Any],
        post_login_network_hint: bool = False,
    ):
        super().__init__()

        self.db = db
        self.warehouse_mgr = warehouse_mgr
        self.material_mgr = material_mgr
        self.dispatch_mgr = dispatch_mgr
        self.current_user = current_user
        self._is_admin = (current_user.get("role") or "").strip().lower() == "admin"

        role_label = "管理员" if self._is_admin else "普通用户"
        uname = current_user.get("username") or ""
        self.setWindowTitle(f"🚚仓库物资调度管理系统 — {uname}（{role_label}）")
        self.resize(1600, 850)

        # ==================== 添加工具栏 ====================

        self.init_ui()
        self._restore_main_window_geometry_from_layout()
        self.refresh_all()          # 初始加载数据
        if post_login_network_hint:
            QTimer.singleShot(400, self._show_post_login_network_hint)

    def _show_post_login_network_hint(self) -> None:
        QMessageBox.information(
            self,
            "配置网络共享",
            "请打开左侧「网络设置」，填写主机共享的 wellsite.db 路径并保存。\n"
            "保存后请重启软件，以便使用主机上的数据、布局与图片记录。\n\n"
            "管理员与普通用户均可在此配置。",
        )

    def init_ui(self):
        # 主分割器（左侧菜单 + 右侧内容）
        main_splitter = QSplitter(Qt.Horizontal)

        # 左侧侧边栏
        self.sidebar = QListWidget()
        sidebar_items = [
            "🏠 首页",
            "🏭 仓库管理",
            "📦 物料管理",
            "🚛 调度管理",
            "🔍 搜索查询",
            "💾 数据备份",
        ]
        if self._is_admin:
            sidebar_items.append("👤 账号分配")
        sidebar_items.append("🌐 网络设置")
        self.sidebar.addItems(sidebar_items)
        self.sidebar.setMaximumWidth(200)
        self.sidebar.setCurrentRow(0)

        # 右侧堆叠页面
        self.stack = QStackedWidget()
        self.page_home = self.create_home_page()
        self.page_warehouse = self.create_warehouse_page()
        self.page_material = self.create_material_page()
        self.page_dispatch = self.create_dispatch_page()
        self.page_search = self.create_search_page()
        self.page_backup = self.create_backup_page()
        if self._is_admin:
            self.page_accounts = self.create_account_page()
        self.page_network = self.create_network_settings_page()
        

        self.stack.addWidget(self.page_home)
        self.stack.addWidget(self.page_warehouse)
        self.stack.addWidget(self.page_material)
        self.stack.addWidget(self.page_dispatch)
        self.stack.addWidget(self.page_search)
        self.stack.addWidget(self.page_backup)
        if self._is_admin:
            self.stack.addWidget(self.page_accounts)
        self.stack.addWidget(self.page_network)
        self.sidebar.currentRowChanged.connect(self.stack.setCurrentIndex)

        main_splitter.addWidget(self.sidebar)
        main_splitter.addWidget(self.stack)
        self.setCentralWidget(main_splitter)

        self.apply_styles()

    def _restore_main_window_geometry_from_layout(self) -> None:
        """若 warehouse_layout.json 含 __main_window__，恢复上次保存的主窗口大小与位置。"""
        scene = getattr(self, "topology_scene", None)
        if scene is None:
            return
        pending = getattr(scene, "pending_main_window_geometry", None)
        if not pending or not isinstance(pending, dict):
            return
        try:
            x = int(pending["x"])
            y = int(pending["y"])
            w = int(pending["width"])
            h = int(pending["height"])
        except (KeyError, TypeError, ValueError):
            return
        if w < 480 or h < 360:
            return
        screen = QGuiApplication.primaryScreen()
        if screen is not None:
            ag = screen.availableGeometry()
            if not ag.intersects(QRect(x, y, w, h)):
                return
        self.setGeometry(x, y, w, h)
        if pending.get("maximized") in (True, "true", 1, "1"):
            self.setWindowState(self.windowState() | Qt.WindowState.WindowMaximized)

    # ====================== 下面保持你原来的代码不变 ======================

    def apply_styles(self):
        self.setStyleSheet("""
            /* 主窗口 */
            QMainWindow {
                background: #f5f7fa;
                color: #303133;
            }

            /* ===== 左侧菜单整体 ===== */
            QListWidget {
                /* 这里改成了渐变色 */
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #2c3e50, stop:1 #34495e);
                /* 右侧用浅色描边，避免与内容区接缝处像黑影 */
                border-right: 1px solid rgba(255, 255, 255, 0.14);
                color: #ecf0f1;
                font-size: 16px;
                outline: none;
            }

            /* 分割条与主内容接缝：去掉默认凸起/阴影感 */
            QSplitter::handle:horizontal {
                background: #f5f7fa;
                width: 2px;
            }
            QSplitter::handle:vertical {
                background: #f5f7fa;
                height: 2px;
            }
            QSplitter::handle:hover {
                background: #e4e7ed;
            }

            /* ===== 每一项 ===== */
            QListWidget::item {
                padding: 14px 20px;
                border: none;
            }

            /* ===== 鼠标悬停 ===== */
            QListWidget::item:hover {
                background: #34495e;
            }

            /* 👇 只有点击才明显 */
            QListWidget::item:selected {
               background: #409eff;
               color: white;
               border-left: 4px solid #409eff;
            }

            /* 分组框 */
            QGroupBox {
                background: white;
                color: #303133;
                font-weight: bold;
                border: 1px solid #dcdfe6;
                border-radius: 8px;
                margin-top: 12px;
                padding: 12px;
            }

            /* 按钮 */
            QPushButton {
                background: #409eff;
                color: white;
                border-radius: 6px;
                padding: 10px 20px;
                font-size: 15px;            
            }
                QPushButton:hover {
                background: #66b1ff;
                           
            }
                           
            /* 输入框 */ 
            QLineEdit, QComboBox, QDateTimeEdit, QSpinBox, QDoubleSpinBox { 
                background: white; 
                color: #303133; 
                border: 1px solid #dcdfe6; 
                border-radius: 6px; 
                padding: 5px; 
                min-height: 28px; 
                font-size: 14px; /* 👈 字体就在这里 */ 
            }
                           
           
            /* 右侧按钮 */
            QComboBox::drop-down {
                width: 30px;   /* 👈 就是这里控制宽度 */  
            }
            /* 箭头 */
            QComboBox::down-arrow {
                    
            }
                           
            QCalendarWidget QWidget { 
                background: #ffffff; 
                color: #303133; 
            }
                           
            QComboBox QAbstractItemView {
                background: white;
                color: #303133;   /* 👈 下拉选项字体颜色 */
                selection-background-color: #409eff;
                selection-color: white;
            }
            
            QSpinBox, QDoubleSpinBox {
                background: white;
                color: #303133;
                border: 1px solid #dcdfe6;
                border-radius: 6px;

                padding: 5px 30px 5px 10px;  /* 👈 很关键 */
                min-height: 28px;
                font-size: 14px;
            }

            QSpinBox::up-button,
            QSpinBox::down-button,
            QDoubleSpinBox::up-button,
            QDoubleSpinBox::down-button {
                width: 25px;
            }
                           
            /* 弹窗背景颜色 */               
            QDialog {
                background: white;
            }

            /* 表格（Segoe UI 在前，保证 Ø、Φ 等拉丁扩展与数字混排；中文回退雅黑） */
            QTableWidget {
                background: white;
                color: #303133;
                border-radius: 6px;
                gridline-color: #ebeef5;
                font-family: "Segoe UI", "Microsoft YaHei", "Arial Unicode MS", "SimHei";
                font-size: 13px;
            }
            QHeaderView::section {
                background: #f5f7fa;
                color: #303133;
                padding: 6px;
                border: none;
                text-align: center;
            }
            /* 行号列与表头交叉的左上角「角块」，不单独设样式时系统常画成深色 */
            QTableCornerButton::section {
                background: #f5f7fa;
                border: none;
            }

            /* 标签 */
            QLabel {
            color: #303133;
            }
       """)

    def create_home_page(self):
        w = QWidget()
        w.setObjectName("homePage")

        bg = asset_path("assets", "home_bg.jpg").replace("\\", "/")
        w.setStyleSheet(
        f"QWidget#homePage {{ border-image: url('{bg}') 0 0 0 0 stretch stretch; }}"
        )

        layout = QVBoxLayout(w)

        label = QLabel()
        label.setAlignment(Qt.AlignCenter)

        # ✨ 用 HTML 分层控制字体大小
        label.setText("""
            <div style="text-align:center;">
               <div style="font-size:60px; font-weight:800; margin-bottom:24px;">
                欢迎使用仓库物资调度管理系统
               </div>
               
               <div style="font-size:15px; opacity:0.8;">
                   开发者：华仁堂-张维
            </div>
            </div>
        """)

        label.setStyleSheet("""
            color: white;
            background: transparent;
            padding: 20px;
        """)

        layout.addStretch()
        layout.addWidget(label, alignment=Qt.AlignCenter)
        layout.addStretch()

        return w

    def create_warehouse_page(self):
        """仓库管理页面 - 已增加删除按钮"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        box = QGroupBox("仓库管理")
        v = QVBoxLayout(box)

        h_layout = QHBoxLayout()
        self.warehouse_input = QLineEdit()
        self.warehouse_input.setPlaceholderText("请输入新仓库名称")
        btn_add = QPushButton("➕ 添加仓库")
        btn_add.clicked.connect(self.add_warehouse)

        h_layout.addWidget(self.warehouse_input)
        h_layout.addWidget(btn_add)

        self.warehouse_list = QListWidget()
        self.warehouse_list.setSelectionMode(QListWidget.SingleSelection)

        btn_delete = QPushButton("🗑 删除选中仓库")
        btn_delete.clicked.connect(self.delete_selected_warehouse)
        btn_delete.setStyleSheet("background: #f56c6c;")
        if not self._is_admin:
            btn_delete.hide()

        v.addLayout(h_layout)
        v.addWidget(self.warehouse_list)
        v.addWidget(btn_delete)

        layout.addWidget(box)
        return widget

    def _ensure_admin(self) -> bool:
        if self._is_admin:
            return True
        QMessageBox.warning(self, "无权限", "仅管理员可进行此操作。")
        return False

    def create_material_page(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        box = QGroupBox("物料管理")
        v = QVBoxLayout(box)

        # ---------- 新增类别（在「删除类别」正上方）----------
        h_cat_add = QHBoxLayout()
        self.category_input = QLineEdit()
        self.category_input.setPlaceholderText("请输入新类别名称")
        btn_add_category = QPushButton("➕ 新增类别")
        btn_add_category.clicked.connect(self.add_category)
        h_cat_add.addWidget(self.category_input, 1)
        h_cat_add.addWidget(btn_add_category)
        v.addLayout(h_cat_add)

        # ---------- 删除类别（仅管理员可见）----------
        cat_del_wrap = QWidget()
        h_cat_del = QHBoxLayout(cat_del_wrap)
        h_cat_del.setContentsMargins(0, 0, 0, 0)
        self.category_delete_combo = QComboBox()
        self.category_delete_combo.setMinimumWidth(160)
        btn_delete_category = QPushButton("🗑 删除类别")
        btn_delete_category.setStyleSheet("background: #f56c6c;")
        btn_delete_category.clicked.connect(self.delete_selected_category)
        h_cat_del.addWidget(self.category_delete_combo, 1)
        h_cat_del.addWidget(btn_delete_category)
        v.addWidget(cat_del_wrap)
        if not self._is_admin:
            cat_del_wrap.hide()

        # ---------- 新增物料 ----------
        h_material = QHBoxLayout()
        self.material_category_combo = QComboBox()
        self.material_category_combo.setPlaceholderText("请选择类别")
        self.material_name = QLineEdit()
        self.material_name.setPlaceholderText("请输入物料名称")
        self.material_model = QLineEdit()
        self.material_model.setPlaceholderText("型号（可选，多个用逗号分隔）")
        self.material_unit = QLineEdit()
        self.material_unit.setPlaceholderText("单位（如：支、袋、件）")
        btn_add_material = QPushButton("➕ 新增物品")
        btn_add_material.clicked.connect(self.add_material)

        h_material.addWidget(self.material_category_combo)
        h_material.addWidget(self.material_name)
        h_material.addWidget(self.material_model)
        h_material.addWidget(self.material_unit)
        h_material.addWidget(btn_add_material)
        v.addLayout(h_material)

        # ---------- 入库按钮 ----------
        btn_add_stock = QPushButton("📥 物料入库")
        btn_add_stock.clicked.connect(self.show_add_stock_dialog)
        v.addWidget(btn_add_stock)

        # ---------- 物料列表 ----------
        self.material_list = QListWidget()
        v.addWidget(self.material_list)
        if self._is_admin:
            self.material_list.setContextMenuPolicy(Qt.CustomContextMenu)
            self.material_list.customContextMenuRequested.connect(self.show_context_menu)
        else:
            self.material_list.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)

        layout.addWidget(box)
        return widget
    
    def show_context_menu(self, pos):
        if not self._ensure_admin():
            return
        item = self.material_list.itemAt(pos)
        if item:
            menu = QMenu()
            edit_action = menu.addAction("编辑")
            del_action = menu.addAction("删除")

            action = menu.exec(self.material_list.mapToGlobal(pos))
            if action is None:
                return

            text = item.text()
            try:
                material_id = int(text.split(" - ")[0])
            except Exception:
                QMessageBox.warning(self, "错误", "解析失败")
                return

            if action == edit_action:
                if not self.material_mgr.get_material(material_id):
                    QMessageBox.warning(self, "错误", "找不到该物料")
                    return
                dialog = EditMaterialDialog(self, self.material_mgr, material_id)
                if dialog.exec() == QDialog.Accepted:
                    self.refresh_material_list()
                return

            if action == del_action:
                reply = QMessageBox.question(
                    self, "确认删除",
                    f"确定删除 {text} 吗？\n\n将同时清除各仓库中该物料的库存记录。"
                    f"若该物料出现在调度历史中，则需先删除相关调度单。",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No,
                )

                if reply == QMessageBox.Yes:
                    try:
                        self.material_mgr.delete_material(material_id)
                        self.refresh_material_list()
                        self.refresh_dispatch_page()
                    except ValueError as e:
                        QMessageBox.warning(self, "无法删除", str(e))
                    except Exception as e:
                        QMessageBox.critical(self, "删除失败", str(e))

    def show_add_stock_dialog(self):
        dialog = AddStockDialog(self, self.warehouse_mgr, self.material_mgr)
        if dialog.exec() == QDialog.Accepted:
           self.refresh_material_list()

    def add_category(self):
        name = self.category_input.text().strip()
        if not name:
            QMessageBox.warning(self, "提示", "类别名称不能为空")
            return
        try:
           self.material_mgr.add_category(name)
           self.category_input.clear()
           QMessageBox.information(self, "成功", f"类别【{name}】添加成功")
           self.refresh_material_list()
        except Exception as e:
           QMessageBox.critical(self, "失败", str(e))

    def delete_selected_category(self):
        if not self._ensure_admin():
            return
        cid = self.category_delete_combo.currentData()
        if cid is None:
            QMessageBox.warning(self, "提示", "暂无类别可删除，请先在下拉框中选择要删除的类别")
            return
        name = self.category_delete_combo.currentText()
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除类别「{name}」吗？\n\n"
            "该类别下的物料将变为「未分类」，物料本身不会被删除。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        try:
            self.material_mgr.delete_category(int(cid))
            self.refresh_material_list()
            QMessageBox.information(self, "成功", f"类别「{name}」已删除")
        except ValueError as e:
            QMessageBox.warning(self, "无法删除", str(e))
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))

    def refresh_material_category_combo(self):
        self.material_category_combo.clear()
        self.category_delete_combo.clear()
        categories = self.material_mgr.list_categories()
        for c in categories:
            self.material_category_combo.addItem(c["name"], c["id"])
            self.category_delete_combo.addItem(c["name"], c["id"])

    def create_dispatch_page(self):
        """调度管理页面 - 支持创建调度记录 + 拓扑图"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # ==================== 顶部工具栏：左「业务+拓扑」、右「数据」====================
        top_layout = QHBoxLayout()
        top_layout.setContentsMargins(4, 4, 4, 10)
        top_layout.setSpacing(0)

        btn_add_dispatch = QPushButton("➕ 创建新调度记录")
        btn_add_dispatch.clicked.connect(self.show_add_dispatch_dialog)
        btn_add_dispatch.setStyleSheet(
            "padding: 10px 18px; font-size: 14px; font-weight: bold;"
        )

        btn_save_layout = QPushButton("💾 保存布局")
        btn_save_layout.clicked.connect(self.save_topology_layout)

        btn_bg = QPushButton("🖼 更换背景")
        btn_bg.clicked.connect(self.change_topology_background)
        btn_bg_clear = QPushButton("🧹 清除背景")
        btn_bg_clear.clicked.connect(self.clear_topology_background)

        btn_refresh = QPushButton("🔄 刷新")
        btn_refresh.clicked.connect(self.refresh_dispatch_page)

        btn_export = QPushButton("📊 导出 Excel")
        btn_export.clicked.connect(self.export_dispatch_excel)

        left_bar = QHBoxLayout()
        left_bar.setSpacing(8)
        left_bar.addWidget(btn_add_dispatch)
        left_bar.addSpacing(20)
        left_bar.addWidget(btn_save_layout)
        left_bar.addWidget(btn_bg)
        left_bar.addWidget(btn_bg_clear)

        right_bar = QHBoxLayout()
        right_bar.setSpacing(8)
        right_bar.addWidget(btn_refresh)
        right_bar.addWidget(btn_export)
        if not self._is_admin:
            btn_export.hide()

        top_layout.addLayout(left_bar)
        top_layout.addStretch(1)
        top_layout.addLayout(right_bar)

        # ==================== 主内容：拓扑图 + 表格 ====================
        self.dispatch_page_splitter = QSplitter(Qt.Vertical)
        splitter = self.dispatch_page_splitter

        # 上方：拓扑图（展示仓库节点）
        # 上方：拓扑图（展示仓库节点）
        self.topology_scene = TopologyScene()
        self.topology_scene.dispatch_mgr = self.dispatch_mgr
        self.topology_scene.load_layout_from_disk()   # 先读取保存的布局和颜色

        warehouses = self.warehouse_mgr.list_warehouses()
        self.topology_scene.build(warehouses)         # 构建节点，优先使用保存数据

        self.topology_view = QGraphicsView(self.topology_scene)
        self.topology_view.setFrameShape(QFrame.Shape.NoFrame)
        self.topology_view.setFrameShadow(QFrame.Shadow.Plain)
        self.topology_view.setStyleSheet("QGraphicsView { border: none; background: #f5f7fa; }")
        self.topology_view.setRenderHint(QPainter.Antialiasing)
        self.topology_view.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)
        self.topology_view.setMinimumHeight(380)

        # 下方：调度记录表格
        self.dispatch_table = QTableWidget(0, 6)
        self.dispatch_table.setHorizontalHeaderLabels(["ID", "调出仓库", "调入仓库", "所含物品", "执行人", "时间"])
        # --- 添加以下设置 ---
        self.dispatch_table.setEditTriggers(QTableWidget.NoEditTriggers) # 禁止编辑
        self.dispatch_table.setSelectionBehavior(QTableWidget.SelectRows) # 点击选中整行
        _tbl_font = QFont()
        _tbl_font.setFamilies(["Segoe UI", "Microsoft YaHei", "Arial Unicode MS", "SimHei"])
        _tbl_font.setPointSize(10)
        self.dispatch_table.setFont(_tbl_font)
        header = self.dispatch_table.horizontalHeader()
        _hdr_font = QFont()
        _hdr_font.setFamilies(["Segoe UI", "Microsoft YaHei", "Arial Unicode MS"])
        _hdr_font.setPointSize(14)
        _hdr_font.setBold(True)
        header.setFont(_hdr_font)
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)  # ⭐核心
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.dispatch_table.itemDoubleClicked.connect(self.show_dispatch_detail_dialog)
        if self._is_admin:
            self.dispatch_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
            self.dispatch_table.customContextMenuRequested.connect(
                self._on_dispatch_table_context_menu
            )
        else:
            self.dispatch_table.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)
        self.dispatch_table.setWordWrap(True)
        self.dispatch_table.resizeRowsToContents()
        #self.dispatch_table.setMaximumHeight(150)  # 👈 改成你想要的高度

        splitter.addWidget(self.topology_view)
        splitter.addWidget(self.dispatch_table)
        pending_ss = getattr(self.topology_scene, "pending_dispatch_splitter_sizes", None)
        if pending_ss and len(pending_ss) >= 2:
            splitter.setSizes([pending_ss[0], pending_ss[1]])
        else:
            splitter.setSizes([420, 150])   # 默认上下比例

        layout.addLayout(top_layout)
        layout.addWidget(splitter)
        return widget

    # ===================== 业务方法 =====================

    def refresh_all(self):
        self.refresh_warehouse_list()
        self.refresh_material_list()
        self.refresh_dispatch_table()
        self.refresh_topology()

    def refresh_warehouse_list(self):
        self.warehouse_list.clear()
        warehouses = self.warehouse_mgr.list_warehouses()
        for w in warehouses:
            item_text = f"{w['id']} - {w['name']}"
            self.warehouse_list.addItem(item_text)

    def refresh_material_list(self):
        self.refresh_material_category_combo()
        self.material_list.clear()
        materials = self.material_mgr.list_materials()
        for m in materials:
            category = m.get("category") or "未分类"
            self.material_list.addItem(f"{m['id']} - [{category}] {m['name']} ({m['unit']})")

    def refresh_topology(self):
       # 保存当前拖拽位置
        self.topology_scene.save_layout_to_disk()
        # 重新构建场景，build 会使用 saved_layout 保持上次拖拽布局
        warehouses = self.warehouse_mgr.list_warehouses()
        self.topology_scene.build(warehouses)

    def refresh_dispatch_table(self):
        query = """
        SELECT r.id,
               w_from.name AS from_name,
               w_to.name AS to_name,
               r.executor,
               r.timestamp
        FROM dispatch_records r
        JOIN warehouses w_from ON r.from_warehouse_id = w_from.id
        JOIN warehouses w_to ON r.to_warehouse_id = w_to.id
        ORDER BY r.timestamp DESC
        LIMIT 100
        """

        records = self.db.fetchall(query)
        ids = [int(r["id"]) for r in records]
        bulk = self.dispatch_mgr.list_dispatch_items_for_records(ids)

        self.dispatch_table.setRowCount(len(records))

        for row, r in enumerate(records):
            self.dispatch_table.setItem(row, 0, QTableWidgetItem(str(r['id'])))
            self.dispatch_table.setItem(row, 1, QTableWidgetItem(r['from_name']))
            self.dispatch_table.setItem(row, 2, QTableWidgetItem(r['to_name']))
            items = bulk.get(int(r["id"]), [])
            mat_raw = DispatchManager.format_items_summary(items) if items else ""
            mat_cell = _format_dispatch_materials_qty_display(mat_raw) if mat_raw else "无"
            self.dispatch_table.setItem(row, 3, QTableWidgetItem(mat_cell))
            self.dispatch_table.setItem(row, 4, QTableWidgetItem(r['executor']))
            self.dispatch_table.setItem(row, 5, QTableWidgetItem(r['timestamp']))
        self.dispatch_table.resizeRowsToContents()

    def add_warehouse(self):
        name = self.warehouse_input.text().strip()
        if not name:
            QMessageBox.warning(self, "提示", "仓库名称不能为空！")
            return
        try:
            self.warehouse_mgr.add_warehouse(name)
            self.warehouse_input.clear()
            self.refresh_warehouse_list()
            self.refresh_topology()
            QMessageBox.information(self, "成功", f"仓库【{name}】添加成功！")
        except Exception as e:
            QMessageBox.critical(self, "添加失败", str(e))

    def delete_selected_warehouse(self):
        if not self._ensure_admin():
            return
        current_item = self.warehouse_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "提示", "请先选择要删除的仓库！")
            return

        text = current_item.text()
        try:
            warehouse_id = int(text.split(" - ")[0])
            warehouse_name = text.split(" - ")[1]
        except:
            QMessageBox.warning(self, "错误", "无法解析仓库信息")
            return

        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除仓库【{warehouse_name}】吗？\n\n如果该仓库已有调度记录，将无法删除！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        try:
            self.warehouse_mgr.delete_warehouse(warehouse_id)
            self.refresh_warehouse_list()
            self.refresh_topology()
            QMessageBox.information(self, "成功", f"仓库【{warehouse_name}】已删除")
        except ValueError as ve:
            QMessageBox.warning(self, "无法删除", str(ve))
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))

    def add_material(self):
        category_id = self.material_category_combo.currentData()
        name = self.material_name.text().strip()
        model = self.material_model.text().strip()
        unit = self.material_unit.text().strip()
        if category_id is None:
            QMessageBox.warning(self, "提示", "请先选择物料类别")
            return
        if not name or not unit:
            QMessageBox.warning(self, "提示", "请填写物料名称和单位")
            return
        try:
            self.material_mgr.add_material(name, unit, model=model or None, category_id=category_id)
            self.material_name.clear()
            self.material_model.clear()
            self.material_unit.clear()
            self.refresh_all()
            QMessageBox.information(self, "成功", f"物料 '{name}' 添加成功")
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))

    def refresh_dispatch_page(self):
        """刷新调度页面（拓扑图 + 表格）"""
        self.refresh_topology()
        self.refresh_dispatch_table()

    def show_add_dispatch_dialog(self):
        """弹出创建调度记录对话框"""
        dialog = AddDispatchDialog(self, self.warehouse_mgr, self.material_mgr, self.dispatch_mgr)
        if dialog.exec() == QDialog.Accepted:
            self.refresh_dispatch_page()   # 创建成功后刷新

    def _on_dispatch_table_context_menu(self, pos):
        idx = self.dispatch_table.indexAt(pos)
        if not idx.isValid():
            return
        row = idx.row()
        menu = QMenu(self)
        act_edit = menu.addAction("编辑")
        chosen = menu.exec(self.dispatch_table.viewport().mapToGlobal(pos))
        if chosen != act_edit:
            return
        it = self.dispatch_table.item(row, 0)
        if not it:
            return
        try:
            rid = int(it.text())
        except ValueError:
            return
        if not self.dispatch_mgr.get_dispatch_detail(rid):
            QMessageBox.warning(self, "提示", "未找到该调度记录")
            return
        dialog = AddDispatchDialog(
            self,
            self.warehouse_mgr,
            self.material_mgr,
            self.dispatch_mgr,
            edit_record_id=rid,
        )
        if dialog.exec() == QDialog.Accepted:
            self.refresh_dispatch_page()

    def change_topology_background(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择背景图片",
            "",
            "Image Files (*.png *.jpg *.jpeg *.bmp *.webp)",
        )
        if not path:
            return
        try:
            stored = copy_file_to_shared(path, "Background images")
        except OSError as e:
            QMessageBox.critical(self, "保存失败", f"无法复制背景图片到本机数据目录：\n{e}")
            return
        self.topology_scene.set_background_image(stored)

    def clear_topology_background(self):
        self.topology_scene.set_background_image("")

    def save_topology_layout(self):
        sp_sizes: Optional[List[int]] = None
        sp = getattr(self, "dispatch_page_splitter", None)
        if sp is not None:
            ss = sp.sizes()
            if isinstance(ss, list) and len(ss) >= 2:
                sp_sizes = [int(ss[0]), int(ss[1])]
        self.topology_scene.save_layout_to_disk(self, dispatch_splitter_sizes=sp_sizes)
        QMessageBox.information(
            self,
            "成功",
            "当前布局、主窗口大小与调度记录列表区域比例已保存，下次启动将自动恢复。",
        )

    def export_dispatch_excel(self):
        """每单占一行：全部物料合并在一格（换行），与早期导出一致。"""
        if not self._is_admin:
            QMessageBox.warning(self, "无权限", "仅管理员可导出调度 Excel。")
            return
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment
        from PySide6.QtWidgets import QFileDialog

        now = datetime.now()
        default_name = f"{now.year}年{now.month}月{now.day}日调度记录.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "调度记录"

        headers = [
            "ID",
            "执行人",
            "调出仓库",
            "调入仓库",
            "时间",
            "全部物料",
            "图片",
        ]
        ws.append(headers)

        wrap_top = Alignment(wrap_text=True, vertical="top")

        records = self.dispatch_mgr.list_records()
        row_index = 2

        for r in records:
            items = self.dispatch_mgr.get_dispatch_items(int(r["id"]))
            mat_text = DispatchManager.format_items_summary(items)
            img_path = r.get("image_path")
            img_ok = bool(img_path and os.path.isfile(img_path))

            ws.cell(row=row_index, column=1, value=r["id"])
            ws.cell(row=row_index, column=2, value=r.get("executor"))
            ws.cell(row=row_index, column=3, value=r.get("from_warehouse"))
            ws.cell(row=row_index, column=4, value=r.get("to_warehouse"))
            ws.cell(row=row_index, column=5, value=r.get("timestamp"))
            c_mat = ws.cell(row=row_index, column=6, value=mat_text)
            c_mat.alignment = wrap_top
            c_img = ws.cell(
                row=row_index,
                column=7,
                value=os.path.basename(img_path) if img_ok else "",
            )
            c_img.alignment = wrap_top

            if img_ok:
                try:
                    img = XLImage(img_path)
                    img.width = 100
                    img.height = 80
                    ws.add_image(img, f"G{row_index}")
                except Exception as e:
                    print("图片插入失败:", e)

            row_index += 1

        dims = {
            "A": 8,
            "B": 10,
            "C": 14,
            "D": 14,
            "E": 18,
            "F": 44,
            "G": 14,
        }
        for col, w in dims.items():
            ws.column_dimensions[col].width = w

        wb.save(path)
    
    def create_search_page(self):
        """搜索查询页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # ---------- 库存搜索 ----------
        stock_group = QGroupBox("库存搜索")
        v1 = QVBoxLayout(stock_group)

        h1 = QHBoxLayout()
        self.search_material_input = QLineEdit()
        self.search_material_input.setPlaceholderText("输入物品名称或型号")
        btn_search_stock = QPushButton("🔍 搜索库存")
        btn_search_stock.clicked.connect(self.search_stock)
        h1.addWidget(self.search_material_input)
        h1.addWidget(btn_search_stock)

        self.stock_table = QTableWidget(0, 3)
        self.stock_table.setHorizontalHeaderLabels(["仓库", "物品", "数量"])
        header = self.stock_table.horizontalHeader()
        header.setFont(QFont("Arial", 14))   # 👈 这里改表头字体大小
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.stock_table.setSelectionBehavior(QTableWidget.SelectRows)

        v1.addLayout(h1)
        v1.addWidget(self.stock_table)

        # ---------- 调度记录查询 ----------
        dispatch_group = QGroupBox("调度记录查询")
        v2 = QVBoxLayout(dispatch_group)

        h2 = QHBoxLayout()
        self.start_date_input = QDateTimeEdit()
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDisplayFormat("yyyy-MM-dd")
        self.start_date_input.setDateTime(QDateTime.currentDateTime().addDays(-7))
        self.start_date_input.setReadOnly(False)
        self.start_date_input.lineEdit().setReadOnly(True)
        self.end_date_input = QDateTimeEdit()
        self.end_date_input.setCalendarPopup(True)
        self.end_date_input.setDisplayFormat("yyyy-MM-dd")
        self.end_date_input.setDateTime(QDateTime.currentDateTime())
        self.end_date_input.setReadOnly(False)
        self.end_date_input.lineEdit().setReadOnly(True)
        btn_search_dispatch = QPushButton("🔍 查询调度")
        btn_search_dispatch.clicked.connect(self.search_dispatch)
        h2.addWidget(self.start_date_input)
        h2.addWidget(self.end_date_input)
        h2.addWidget(btn_search_dispatch)

        self.dispatch_table_search = QTableWidget(0, 6)
        self.dispatch_table_search.setHorizontalHeaderLabels(
            ["ID", "调出仓库", "调入仓库", "所含物品", "执行人", "时间"]
        )
        header2 = self.dispatch_table_search.horizontalHeader()
        header2.setFont(QFont("Microsoft YaHei", 14))
        header2.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        header2.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header2.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header2.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header2.setSectionResizeMode(3, QHeaderView.Stretch)
        header2.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header2.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.dispatch_table_search.setEditTriggers(QTableWidget.NoEditTriggers)
        self.dispatch_table_search.setSelectionBehavior(QTableWidget.SelectRows)
        _search_tbl_font = QFont()
        _search_tbl_font.setFamilies(["Segoe UI", "Microsoft YaHei", "Arial Unicode MS", "SimHei"])
        _search_tbl_font.setPointSize(10)
        self.dispatch_table_search.setFont(_search_tbl_font)
        self.dispatch_table_search.setWordWrap(True)

        v2.addLayout(h2)
        v2.addWidget(self.dispatch_table_search)

        layout.addWidget(stock_group)
        layout.addWidget(dispatch_group)

        return widget
    

    def search_stock(self):
        keyword = self.search_material_input.text().strip()
        if not keyword:
            QMessageBox.warning(self, "提示", "请输入物品名称或型号")
            return

        # 查询仓库物料库存
        # 假设 warehouse_mgr 有方法 list_warehouse_items_by_keyword(keyword)
        results = self.warehouse_mgr.search_material(keyword)  # 返回 [{'warehouse':'仓库A','name':'物品','quantity':10},...]

        self.stock_table.setRowCount(len(results))
        if not results:
            QMessageBox.information(self, "查询结果", "没有库存或名称不对")
            return
        for row, r in enumerate(results):
            self.stock_table.setItem(row, 0, QTableWidgetItem(r['warehouse']))
            self.stock_table.setItem(row, 1, QTableWidgetItem(r['name']))
            try:
                qv = float(r["quantity"])
                qtxt = str(int(qv)) if qv == int(qv) else str(qv).rstrip("0").rstrip(".")
            except (TypeError, ValueError):
                qtxt = str(r["quantity"])
            self.stock_table.setItem(row, 2, QTableWidgetItem(qtxt))
    
    def search_dispatch(self):
        start = self.start_date_input.date().toString("yyyy-MM-dd")
        end = self.end_date_input.date().toString("yyyy-MM-dd")

        if self.start_date_input.dateTime() > self.end_date_input.dateTime():
           QMessageBox.warning(self, "提示", "开始时间不能晚于结束时间")
           return
        
        try:
            query = """
            SELECT r.id,
                   w_from.name AS from_name,
                   w_to.name AS to_name,
                   r.executor,
                   r.timestamp
            FROM dispatch_records r
            JOIN warehouses w_from ON r.from_warehouse_id = w_from.id
            JOIN warehouses w_to ON r.to_warehouse_id = w_to.id
            WHERE date(r.timestamp) BETWEEN ? AND ?
            ORDER BY r.timestamp DESC
            """
            records = self.db.fetchall(query, (start, end))
            ids = [int(r["id"]) for r in records]
            bulk = self.dispatch_mgr.list_dispatch_items_for_records(ids)

            self.dispatch_table_search.setRowCount(len(records))
            if not records:
                QMessageBox.information(self, "查询结果", "没有符合条件的调度记录")
                return
            for row, r in enumerate(records):
                self.dispatch_table_search.setItem(row, 0, QTableWidgetItem(str(r['id'])))
                self.dispatch_table_search.setItem(row, 1, QTableWidgetItem(r['from_name']))
                self.dispatch_table_search.setItem(row, 2, QTableWidgetItem(r['to_name']))
                items = bulk.get(int(r["id"]), [])
                mat_raw = DispatchManager.format_items_summary(items) if items else ""
                mat_cell = _format_dispatch_materials_qty_display(mat_raw) if mat_raw else "无"
                self.dispatch_table_search.setItem(row, 3, QTableWidgetItem(mat_cell))
                self.dispatch_table_search.setItem(row, 4, QTableWidgetItem(r['executor']))
                self.dispatch_table_search.setItem(row, 5, QTableWidgetItem(r['timestamp']))
            self.dispatch_table_search.resizeRowsToContents()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"查询失败: {e}")    
    

    def _read_backup_ui_settings(self) -> dict:
        path = backup_ui_settings_file()
        if not path or not os.path.isfile(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _write_backup_ui_settings(self, data: dict) -> None:
        path = backup_ui_settings_file()
        if not path:
            return
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _apply_backup_page_background(self, widget: QWidget) -> None:
        """刷新备份页背景（由 BackupPageRoot 绘制，不依赖 QSS url）。"""
        if isinstance(widget, BackupPageRoot):
            widget.reload_background(self)

    def _pick_backup_page_background(self, page: QWidget) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择备份页背景图片",
            shared_background_dir() if os.path.isdir(shared_background_dir()) else user_data_root(),
            "Image Files (*.png *.jpg *.jpeg *.bmp *.webp);;All Files (*.*)",
        )
        if not path:
            return
        norm = os.path.normpath(path)
        if not os.path.isfile(norm):
            QMessageBox.warning(self, "无效路径", f"找不到文件：\n{norm}")
            return
        probe = QPixmap(norm)
        if probe.isNull():
            QMessageBox.warning(
                self,
                "无法作为图片加载",
                "该路径存在，但 Qt 无法解码为图片（可能格式不支持或文件损坏）。\n"
                "请换用 PNG / JPG 等常见格式。",
            )
            return
        try:
            stored = copy_file_to_shared(norm, "Background images")
        except OSError as e:
            QMessageBox.critical(self, "保存失败", f"无法复制背景图片到本机数据目录：\n{e}")
            return
        data = self._read_backup_ui_settings()
        data["background_image"] = stored
        self._write_backup_ui_settings(data)
        self._apply_backup_page_background(page)
        QMessageBox.information(self, "成功", "备份页背景已更新。\n图片与 backup_ui_settings.json 已保存在本机 exe 旁。")

    def _clear_backup_page_background(self, page: QWidget) -> None:
        data = self._read_backup_ui_settings()
        data["background_image"] = ""
        self._write_backup_ui_settings(data)
        self._apply_backup_page_background(page)
        QMessageBox.information(self, "成功", "已清除自定义背景，将使用默认图或渐变。")

    def create_backup_page(self):
        """数据备份页面（可自选背景图，配置见 backup_ui_settings.json）"""
        widget = BackupPageRoot(self)
        self._apply_backup_page_background(widget)

        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 16, 20, 24)
        layout.setSpacing(18)

        top_bar = QHBoxLayout()
        top_bar.setSpacing(12)
        btn_backup = QPushButton("立即备份")
        btn_backup.setObjectName("backupMainBtn")
        btn_backup.setMinimumWidth(160)
        btn_backup.setMinimumHeight(44)
        btn_backup.clicked.connect(self.backup_data)
        btn_bg = QPushButton("设置背景")
        btn_bg.setObjectName("backupOutlineBtn")
        btn_bg_clear = QPushButton("清除背景")
        btn_bg_clear.setObjectName("backupOutlineBtn")
        btn_bg.clicked.connect(lambda: self._pick_backup_page_background(widget))
        btn_bg_clear.clicked.connect(lambda: self._clear_backup_page_background(widget))
        top_bar.addStretch()
        top_bar.addWidget(btn_backup)
        top_bar.addWidget(btn_bg)
        top_bar.addWidget(btn_bg_clear)
        top_bar.addStretch()

        title = QLabel("💾 数据备份")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet(
            "font-size: 22px; font-weight: bold; color: #303133; background: transparent;"
        )
        hint = QLabel(
            "「立即备份」将仓库、物料、调度记录导出为 JSON；"
            "「设置背景 / 清除背景」可更换本页背景图。"
        )
        hint.setAlignment(Qt.AlignCenter)
        hint.setWordWrap(True)
        hint.setStyleSheet("font-size: 15px; color: #606266; background: transparent;")

        tip_bg = QLabel("背景路径保存在本机 exe 旁的 backup_ui_settings.json，可随时更换。")
        tip_bg.setAlignment(Qt.AlignCenter)
        tip_bg.setWordWrap(True)
        tip_bg.setStyleSheet("font-size: 12px; color: #909399; background: transparent;")

        layout.addLayout(top_bar)
        layout.addWidget(title)
        layout.addWidget(hint)
        layout.addWidget(tip_bg)
        layout.addStretch()

        return widget

    def create_account_page(self):
        """管理员账号分配：新增账号、删除账号。"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(28, 24, 28, 28)
        layout.setSpacing(16)

        title = QLabel("👤 账号分配")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #303133;")
        layout.addWidget(title)

        add_box = QGroupBox("添加账号")
        add_layout = QVBoxLayout(add_box)
        form = QFormLayout()
        self.account_username_edit = QLineEdit()
        self.account_username_edit.setPlaceholderText("请输入用户名")
        self.account_password_edit = QLineEdit()
        self.account_password_edit.setPlaceholderText("请输入密码")
        self.account_password_edit.setEchoMode(QLineEdit.Password)
        self.account_role_combo = QComboBox()
        self.account_role_combo.addItem("普通用户", "user")
        self.account_role_combo.addItem("管理员", "admin")
        form.addRow("用户名：", self.account_username_edit)
        form.addRow("密码：", self.account_password_edit)
        form.addRow("角色：", self.account_role_combo)
        add_layout.addLayout(form)

        add_row = QHBoxLayout()
        add_row.addStretch()
        btn_add_account = QPushButton("添加账号")
        btn_add_account.clicked.connect(self.add_account)
        add_row.addWidget(btn_add_account)
        add_layout.addLayout(add_row)
        layout.addWidget(add_box)

        list_box = QGroupBox("账号列表")
        list_layout = QVBoxLayout(list_box)
        self.account_table = QTableWidget(0, 3)
        self.account_table.setHorizontalHeaderLabels(["ID", "用户名", "角色"])
        self.account_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.account_table.setSelectionMode(QTableWidget.SingleSelection)
        self.account_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.account_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.account_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.account_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        list_layout.addWidget(self.account_table)

        list_btn_row = QHBoxLayout()
        btn_refresh = QPushButton("刷新列表")
        btn_delete = QPushButton("删除选中账号")
        btn_delete.setStyleSheet("background: #f56c6c;")
        btn_refresh.clicked.connect(self.load_accounts)
        btn_delete.clicked.connect(self.delete_selected_account)
        list_btn_row.addStretch()
        list_btn_row.addWidget(btn_refresh)
        list_btn_row.addWidget(btn_delete)
        list_layout.addLayout(list_btn_row)
        layout.addWidget(list_box, stretch=1)

        tip = QLabel("提示：删除账号不会删除该账号过去创建的调度记录。为了安全，不能删除当前登录账号，也不能删除最后一个管理员。")
        tip.setWordWrap(True)
        tip.setStyleSheet("color: #606266; font-size: 13px;")
        layout.addWidget(tip)

        self.load_accounts()
        return widget

    def load_accounts(self) -> None:
        if not hasattr(self, "account_table"):
            return
        rows = self.db.fetchall(
            """
            SELECT id, username, role
            FROM users
            ORDER BY LOWER(username)
            """
        )
        self.account_table.setRowCount(0)
        for row_data in rows:
            row = self.account_table.rowCount()
            self.account_table.insertRow(row)
            role = (row_data["role"] or "user").strip().lower()
            role_label = "管理员" if role == "admin" else "普通用户"
            values = [str(row_data["id"]), row_data["username"], role_label]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setData(Qt.UserRole, int(row_data["id"]))
                self.account_table.setItem(row, col, item)

    def add_account(self) -> None:
        if not self._ensure_admin():
            return
        username = self.account_username_edit.text().strip()
        password = self.account_password_edit.text()
        role = self.account_role_combo.currentData() or "user"
        if not username or not password:
            QMessageBox.warning(self, "提示", "请输入用户名和密码。")
            return
        if len(password) < 4:
            QMessageBox.warning(self, "提示", "密码至少 4 位。")
            return
        try:
            self.db.execute(
                "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                (username, hash_password(password), role),
            )
        except RuntimeError as e:
            msg = str(e)
            if "UNIQUE" in msg.upper():
                QMessageBox.warning(self, "添加失败", "用户名已存在，请换一个用户名。")
            else:
                QMessageBox.critical(self, "添加失败", msg)
            return
        except Exception as e:
            QMessageBox.critical(self, "添加失败", str(e))
            return

        self.account_username_edit.clear()
        self.account_password_edit.clear()
        self.account_role_combo.setCurrentIndex(0)
        self.load_accounts()
        QMessageBox.information(self, "成功", f"账号「{username}」已添加。")

    def delete_selected_account(self) -> None:
        if not self._ensure_admin():
            return
        row = self.account_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择要删除的账号。")
            return
        id_item = self.account_table.item(row, 0)
        name_item = self.account_table.item(row, 1)
        role_item = self.account_table.item(row, 2)
        if id_item is None or name_item is None or role_item is None:
            QMessageBox.warning(self, "提示", "无法读取选中的账号。")
            return
        user_id = int(id_item.text())
        username = name_item.text()
        role_label = role_item.text()
        if user_id == int(self.current_user.get("id") or 0):
            QMessageBox.warning(self, "无法删除", "不能删除当前正在登录的账号。")
            return
        if role_label == "管理员":
            admin_count = self.db.fetch_scalar(
                "SELECT COUNT(*) FROM users WHERE LOWER(role) = 'admin'"
            )
            if int(admin_count or 0) <= 1:
                QMessageBox.warning(self, "无法删除", "不能删除最后一个管理员账号。")
                return

        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除账号「{username}」吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        try:
            self.db.execute("DELETE FROM users WHERE id = ?", (user_id,))
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e))
            return
        self.load_accounts()
        QMessageBox.information(self, "成功", f"账号「{username}」已删除。")

    def _read_network_settings(self) -> dict:
        if not os.path.isfile(NETWORK_SETTINGS_FILE):
            return {}
        try:
            with open(NETWORK_SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _write_network_settings(self, data: dict) -> None:
        payload = dict(data)
        payload["client"] = True
        with open(NETWORK_SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def _configured_network_db_path(self) -> str:
        data = self._read_network_settings()
        if data.get("enabled") is False:
            return ""
        return str(data.get("db_path") or "").strip()

    def _refresh_network_settings_labels(self) -> None:
        current = os.path.abspath(os.path.normpath(getattr(self.db, "db_path", "")))
        saved = self._configured_network_db_path()
        if hasattr(self, "network_current_db_label"):
            self.network_current_db_label.setText(current)
        if hasattr(self, "network_saved_db_label"):
            if saved:
                lab = saved
            else:
                lab = "仅使用共享数据库（见 network_settings.json）"
            self.network_saved_db_label.setText(lab)
        if hasattr(self, "network_db_path_edit") and saved:
            self.network_db_path_edit.setText(saved)

    def create_network_settings_page(self):
        """网络数据库设置：保存共享 wellsite.db 路径，下次启动时生效。"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(28, 24, 28, 28)
        layout.setSpacing(16)

        title = QLabel("🌐 网络设置")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #303133;")
        layout.addWidget(title)

        current_box = QGroupBox("当前数据库")
        current_layout = QFormLayout(current_box)
        self.network_current_db_label = QLabel("")
        self.network_current_db_label.setWordWrap(True)
        self.network_current_db_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.network_saved_db_label = QLabel("")
        self.network_saved_db_label.setWordWrap(True)
        self.network_saved_db_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        current_layout.addRow("本次正在使用：", self.network_current_db_label)
        current_layout.addRow("已保存的网络路径：", self.network_saved_db_label)
        layout.addWidget(current_box)

        setting_box = QGroupBox("设置共享数据库路径")
        setting_layout = QVBoxLayout(setting_box)
        form = QFormLayout()
        self.network_db_path_edit = QLineEdit()
        self.network_db_path_edit.setPlaceholderText(r"例如：\\192.168.1.20\wellsite_share\wellsite.db")
        form.addRow("共享数据库：", self.network_db_path_edit)
        setting_layout.addLayout(form)

        btn_row = QHBoxLayout()
        btn_browse = QPushButton("选择数据库文件")
        btn_save = QPushButton("保存设置")
        btn_copy = QPushButton("复制当前库到此路径并保存")
        btn_clear = QPushButton("关闭网络设置")
        btn_clear.setStyleSheet("background: #909399;")
        btn_browse.clicked.connect(self.pick_network_db_path)
        btn_save.clicked.connect(self.save_network_settings)
        btn_copy.clicked.connect(self.copy_current_db_to_network_path)
        btn_clear.clicked.connect(self.clear_network_settings)
        btn_row.addWidget(btn_browse)
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_copy)
        btn_row.addWidget(btn_clear)
        setting_layout.addLayout(btn_row)
        btn_clear.hide()
        btn_copy.hide()
        layout.addWidget(setting_box)

        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setMinimumHeight(190)
        help_text.setText(
            "使用方法：\n"
            "1. 在主机上建立共享文件夹（例如 D:\\wellsite_share），并保证本机对该共享有读写权限。\n"
            "2. 在此填写主机上的 wellsite.db 路径，例如 \\\\192.168.1.20\\wellsite_share\\wellsite.db。\n"
            "3. 共享路径下必须已存在主机创建好的 wellsite.db；本程序不会把本机库复制到共享（请在主机端维护库）。\n"
            "4. 保存后需重启软件才会使用新的共享路径。\n\n"
            "启用后，仅在共享路径与 wellsite.db 同目录读写：wellsite.db 与调度附图目录 Picture record。\n"
            "仓库布局 warehouse_layout.json、备份页配置 backup_ui_settings.json、拓扑/备份页背景图 Background images 均在本机 exe 旁，各客户端独立。\n"
            "不能在此关闭网络改单机；若已保存共享路径但主机不可达，程序将无法启动。\n"
            "首次无配置时会弹出「连接共享数据库」向导。\n\n"
            "提醒：主机宜固定 IP；SQLite 适合少量人同时使用。"
        )
        layout.addWidget(help_text)
        layout.addStretch()

        saved = self._configured_network_db_path()
        if saved:
            self.network_db_path_edit.setText(saved)
        self._refresh_network_settings_labels()
        return widget

    def pick_network_db_path(self) -> None:
        initial = self.network_db_path_edit.text().strip() or shared_data_root() or user_data_root()
        if os.path.isfile(initial):
            initial = os.path.dirname(initial)
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择共享数据库 wellsite.db",
            initial,
            "SQLite Database (*.db);;All Files (*.*)",
        )
        if path:
            self.network_db_path_edit.setText(os.path.normpath(path))

    def _network_path_from_edit(self) -> str:
        raw = self.network_db_path_edit.text().strip()
        if not raw:
            return ""
        return os.path.normpath(os.path.expandvars(raw))

    def save_network_settings(self) -> None:
        path = self._network_path_from_edit()
        if not path:
            QMessageBox.warning(self, "提示", "请填写共享数据库路径。")
            return
        parent = os.path.dirname(path)
        if parent and not os.path.isdir(parent):
            QMessageBox.warning(self, "路径无效", f"找不到共享文件夹：\n{parent}")
            return
        if not os.path.isfile(path):
            QMessageBox.warning(
                self,
                "数据库不存在",
                "共享路径下还没有 wellsite.db。\n\n"
                "请在主机上点击“复制当前库到此路径并保存”，客户端只连接已有的主机数据库。",
            )
            return
        self._write_network_settings({"enabled": True, "db_path": path})
        self._refresh_network_settings_labels()
        QMessageBox.information(self, "已保存", "网络数据库路径已保存。\n请重启软件后生效。")

    def clear_network_settings(self) -> None:
        QMessageBox.warning(
            self,
            "提示",
            "本程序仅允许使用共享数据库，不能关闭网络设置。",
        )

    def _copy_file_to_target_root(self, src_path: str, target_root: str, subdir: str) -> str:
        ap = os.path.abspath(os.path.normpath(src_path))
        dest_root = os.path.abspath(os.path.join(target_root, subdir))
        os.makedirs(dest_root, exist_ok=True)
        base = os.path.basename(ap)
        safe = re.sub(r"[^\w\-\.()\u4e00-\u9fff]", "_", base).strip("._") or "file"
        stem, ext = os.path.splitext(safe)
        dest = os.path.join(dest_root, safe)
        n = 0
        while os.path.exists(dest):
            n += 1
            dest = os.path.join(dest_root, f"{stem}_{n}{ext}")
        shutil.copy2(ap, dest)
        return os.path.abspath(dest)

    def _copy_support_files_to_network_root(
        self, dest_db_path: str, target_root: str, src_db_path: str
    ) -> None:
        """将调度附图迁到目标共享根下的 Picture record；布局与备份页 JSON 保留在本机。"""
        os.makedirs(target_root, exist_ok=True)

        copied_photo_paths: Dict[str, str] = {}
        src_parent = os.path.dirname(os.path.abspath(os.path.normpath(src_db_path)))
        photo_dirs: List[str] = []
        if src_parent:
            photo_dirs.append(os.path.join(src_parent, "Picture record"))
        photo_dirs.append(os.path.join(user_data_root(), "Picture record"))
        seen: set[str] = set()
        for root_dir in photo_dirs:
            rd = os.path.abspath(os.path.normpath(root_dir))
            if not rd or rd in seen:
                continue
            seen.add(rd)
            if not os.path.isdir(rd):
                continue
            for name in os.listdir(rd):
                src = os.path.join(rd, name)
                if not os.path.isfile(src):
                    continue
                copied = self._copy_file_to_target_root(src, target_root, "Picture record")
                copied_photo_paths[os.path.normcase(os.path.abspath(src))] = copied

        # 数据库里保存的是图片绝对路径，迁移时同步改成目标共享目录下的路径。
        with sqlite3.connect(dest_db_path) as conn:
            cur = conn.execute(
                "SELECT id, image_path FROM dispatch_records "
                "WHERE image_path IS NOT NULL AND TRIM(image_path) <> ''"
            )
            updates = []
            for rid, img_path in cur.fetchall():
                raw = str(img_path or "").strip()
                if not raw or not os.path.isfile(raw):
                    continue
                copied = copied_photo_paths.get(
                    os.path.normcase(os.path.abspath(os.path.normpath(raw)))
                )
                if not copied:
                    copied = self._copy_file_to_target_root(raw, target_root, "Picture record")
                updates.append((copied, int(rid)))
            if updates:
                conn.executemany(
                    "UPDATE dispatch_records SET image_path = ? WHERE id = ?",
                    updates,
                )
                conn.commit()

    def copy_current_db_to_network_path(self) -> None:
        dest = self._network_path_from_edit()
        if not dest:
            QMessageBox.warning(self, "提示", "请先填写要保存到共享文件夹的数据库路径。")
            return
        parent = os.path.dirname(dest)
        if parent and not os.path.isdir(parent):
            QMessageBox.warning(self, "路径无效", f"找不到共享文件夹：\n{parent}")
            return

        src = os.path.abspath(os.path.normpath(getattr(self.db, "db_path", "")))
        dest_abs = os.path.abspath(os.path.normpath(dest))
        if os.path.normcase(src) == os.path.normcase(dest_abs):
            QMessageBox.information(self, "提示", "当前已经在使用这个数据库路径。")
            return
        if os.path.exists(dest_abs):
            reply = QMessageBox.question(
                self,
                "确认覆盖",
                f"目标文件已存在，是否覆盖？\n{dest_abs}",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        try:
            with sqlite3.connect(dest_abs) as dest_conn:
                self.db.conn.backup(dest_conn)
            self._copy_support_files_to_network_root(dest_abs, os.path.dirname(dest_abs), src)
        except Exception as e:
            QMessageBox.critical(self, "复制失败", f"无法复制数据库或 Picture record：\n{e}")
            return

        self._write_network_settings({"enabled": True, "db_path": dest_abs})
        self._refresh_network_settings_labels()
        QMessageBox.information(
            self,
            "已完成",
            "当前数据库与 Picture record 已复制到共享路径；布局与备份页背景仍在本机 exe 旁。\n请重启软件后生效。",
        )
    
    def backup_data(self):
        """导出系统数据为 JSON 文件"""
        path, _ = QFileDialog.getSaveFileName(
            self, "选择备份文件保存路径", "backup.json", "JSON Files (*.json)"
        )
        if not path:
            return

        try:
            data = {
                "warehouses": self.warehouse_mgr.list_warehouses(),
                "materials": self.material_mgr.list_materials(),
                "dispatch_records": self.dispatch_mgr.list_records()
            }

            with open(path, "w", encoding="utf-8") as f:
               json.dump(data, f, ensure_ascii=False, indent=2)

            QMessageBox.information(self, "备份成功", f"系统数据已成功备份到:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "备份失败", str(e))

    def show_dispatch_detail_dialog(self, item):
        row = item.row()
        record_id = self.dispatch_table.item(row, 0).text()
        
        # 构造数据传递给弹窗
        data = {
            'id': record_id,
            'from': self.dispatch_table.item(row, 1).text(),
            'to': self.dispatch_table.item(row, 2).text(),
            'materials': self.dispatch_table.item(row, 3).text(),
            'executor': self.dispatch_table.item(row, 4).text(),
            'time': self.dispatch_table.item(row, 5).text()
        }
        
        dialog = DispatchDetailDialog(data, self)
        dialog.exec()


class EditMaterialDialog(QDialog):
    """编辑已有物料：名称、单位、类别、型号（逗号分隔，可增删）"""

    def __init__(self, parent, material_mgr: MaterialManager, material_id: int):
        super().__init__(parent)
        self.material_mgr = material_mgr
        self.material_id = material_id
        self.setWindowTitle("编辑物料")
        self.resize(520, 340)

        m = material_mgr.get_material(material_id)
        assert m is not None

        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(10)

        self.name_edit = QLineEdit(m.get("name") or "")
        self.unit_edit = QLineEdit(m.get("unit") or "")
        self.category_combo = QComboBox()
        self.model_edit = QLineEdit()
        raw_models = (m.get("model") or "").strip()
        if raw_models:
            seen = []
            for p in raw_models.split(","):
                t = p.strip()
                if t and t not in seen:
                    seen.append(t)
            self.model_edit.setText(", ".join(seen))

        form.addRow("物料名称：", self.name_edit)
        form.addRow("单位：", self.unit_edit)
        form.addRow("类别：", self.category_combo)
        form.addRow("型号：", self.model_edit)
        layout.addLayout(form)

        tip = QLabel(
            "多个型号用英文逗号分隔；删掉不需要的型号后保存即可移除。"
            "\n型号全部留空表示该物料不设型号（与入库时的空型号一致）。"
        )
        tip.setWordWrap(True)
        tip.setStyleSheet("color: #606266; font-size: 13px;")
        layout.addWidget(tip)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("取消")
        btn_save = QPushButton("保存")
        btn_save.setDefault(True)
        btn_save.setStyleSheet("background: #67c23a; color: white; padding: 8px 20px;")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self.save)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_save)
        layout.addLayout(btn_row)

        self._load_categories(m.get("category_id"))

    def _load_categories(self, current_category_id):
        self.category_combo.clear()
        self.category_combo.addItem("未分类", None)
        for c in self.material_mgr.list_categories():
            self.category_combo.addItem(c["name"], c["id"])

        target = current_category_id
        for i in range(self.category_combo.count()):
            data = self.category_combo.itemData(i)
            if data == target or (data is None and target is None):
                self.category_combo.setCurrentIndex(i)
                return
        self.category_combo.setCurrentIndex(0)

    def save(self):
        name = self.name_edit.text().strip()
        unit = self.unit_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "提示", "物料名称不能为空")
            return
        if not unit:
            QMessageBox.warning(self, "提示", "单位不能为空")
            return

        model_text = self.model_edit.text().strip()
        category_id = self.category_combo.currentData()

        try:
            self.material_mgr.update_material(
                self.material_id,
                name=name,
                unit=unit,
                model=model_text,
                category_id=category_id,
            )
        except ValueError as e:
            QMessageBox.warning(self, "无法保存", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))
            return

        QMessageBox.information(self, "成功", "物料已更新")
        self.accept()


class AddDispatchDialog(QDialog):
    """创建 / 编辑调度记录对话框"""

    def __init__(
        self,
        parent,
        warehouse_mgr: WarehouseManager,
        material_mgr: MaterialManager,
        dispatch_mgr: DispatchManager,
        edit_record_id: Optional[int] = None,
    ):
        super().__init__(parent)
        self.warehouse_mgr = warehouse_mgr
        self.material_mgr = material_mgr
        self.dispatch_mgr = dispatch_mgr
        self.edit_record_id = edit_record_id
        self.image_path = None
        self.setWindowTitle(
            f"编辑调度记录 #{edit_record_id}" if edit_record_id else "创建新调度记录"
        )
        self.resize(960, 700)
        self.setModal(True)

        self.items = []
        self._materials = []

        self.init_ui()
        self.load_warehouses()
        self.load_categories()
        self.load_materials()
        if edit_record_id:
            self._load_dispatch_for_edit(edit_record_id)

    def init_ui(self):

        layout = QVBoxLayout(self)

        # ==================== 基本信息 ====================
        form = QFormLayout()
        form.setSpacing(12)

        self.from_combo = QComboBox()
        self.to_combo = QComboBox()
        self.executor_edit = QLineEdit("张维")
        self.remarks_edit = QLineEdit()
        self.image_path = None

        form.addRow("调出仓库：", self.from_combo)
        form.addRow("调入仓库：", self.to_combo)
        form.addRow("执行人：", self.executor_edit)
        form.addRow("备注：", self.remarks_edit)

        layout.addLayout(form)


        # 图片上传
        img_layout = QHBoxLayout()

        self.img_label = QLabel("未选择图片")
        btn_upload = QPushButton("📷 上传图片")

        btn_upload.clicked.connect(self.select_image)

        img_layout.addWidget(self.img_label)
        img_layout.addWidget(btn_upload)

        layout.addLayout(img_layout)

        # ==================== 物料明细 ====================
        material_group = QVBoxLayout()

        label = QLabel("调度物料明细")
        label.setStyleSheet("font-weight: bold; font-size: 14px;")
        material_group.addWidget(label)

        # ----------------- 物料输入控件 -----------------
        add_layout = QHBoxLayout()

        # 类别（可输入+下拉）
        self.category_combo = QComboBox()
        add_layout.addWidget(QLabel("类别："))
        add_layout.addWidget(self.category_combo, 2)

        # 物品
        self.material_combo = QComboBox()
        self.material_combo.setEditable(True)
        add_layout.addWidget(QLabel("物品："))
        add_layout.addWidget(self.material_combo, 4)

        # 型号
        self.model_combo = QComboBox()
        self.model_combo.setEditable(True)
        add_layout.addWidget(QLabel("型号："))
        add_layout.addWidget(self.model_combo, 3)

        # 数量
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 9999)
        add_layout.addWidget(QLabel("数量："))
        add_layout.addWidget(self.quantity_spin, 2)

        # 按钮
        btn_add_item = QPushButton("➕ 添加")
        btn_add_item.clicked.connect(self.add_material_item)
        add_layout.addWidget(btn_add_item)

        self.category_combo.currentTextChanged.connect(self.load_materials)
        self.material_combo.currentIndexChanged.connect(self.load_models_for_current_material)


        material_group.addLayout(add_layout)

        # ------------------ 调度物料明细表格初始化 ------------------
        self.item_table = QTableWidget(0, 5)
        self.item_table.setHorizontalHeaderLabels(["类别", "物品", "型号", "数量", "操作"])
        self.item_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

        self.item_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.item_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.item_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.item_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.item_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        material_group.addWidget(self.item_table)

        layout.addLayout(material_group)

        # ==================== 底部按钮 ====================
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_cancel = QPushButton("取消")
        btn_save = QPushButton("✅ 保存调度记录")
        btn_save.setStyleSheet("background: #67c23a; color: white; padding: 8px 20px;")

        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self.save_dispatch)

        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)

        layout.addLayout(btn_layout)

    def _load_dispatch_for_edit(self, record_id: int) -> None:
        detail = self.dispatch_mgr.get_dispatch_detail(record_id)
        if not detail:
            return

        def _combo_idx(combo: QComboBox, wid: int) -> int:
            for i in range(combo.count()):
                if combo.itemData(i) == wid:
                    return i
            return -1

        fi = _combo_idx(self.from_combo, int(detail["from_warehouse_id"]))
        if fi >= 0:
            self.from_combo.setCurrentIndex(fi)
        ti = _combo_idx(self.to_combo, int(detail["to_warehouse_id"]))
        if ti >= 0:
            self.to_combo.setCurrentIndex(ti)

        self.executor_edit.setText((detail.get("executor") or "").strip())
        self.remarks_edit.setText((detail.get("remarks") or "").strip())

        ip = (detail.get("image_path") or "").strip()
        if ip and os.path.isfile(ip):
            self.image_path = ip
            self.img_label.setText(os.path.basename(ip))
        else:
            self.image_path = None
            self.img_label.setText("未选择图片")

        self.item_table.setRowCount(0)
        self.items.clear()

        for it in detail.get("items") or []:
            mid = int(it["material_id"])
            m = self.material_mgr.get_material(mid)
            category = "未分类"
            name = (it.get("material_name") or "").strip()
            if m:
                name = (m.get("name") or name).strip()
                cid = m.get("category_id")
                if cid is not None:
                    for c in self.material_mgr.list_categories():
                        if c.get("id") == cid:
                            category = (c.get("name") or "未分类").strip()
                            break

            model_text = (it.get("model_name") or "").strip()
            qty_f = float(it.get("quantity") or 0)
            qshow = (
                str(int(qty_f))
                if qty_f == int(qty_f)
                else str(qty_f).rstrip("0").rstrip(".")
            )

            mid_model = it.get("model_id")
            raw_lab = (it.get("dispatch_model_label") or "").strip()
            if mid_model is not None:
                m_label = None
            else:
                m_label = raw_lab if raw_lab else (model_text or None)

            row = self.item_table.rowCount()
            self.item_table.insertRow(row)
            self.item_table.setItem(row, 0, QTableWidgetItem(category))
            self.item_table.setItem(row, 1, QTableWidgetItem(name))
            self.item_table.setItem(row, 2, QTableWidgetItem(model_text))
            self.item_table.setItem(row, 3, QTableWidgetItem(qshow))
            btn_del = QPushButton("删除")
            btn_del.clicked.connect(lambda _, r=row: self.delete_item_row(r))
            self.item_table.setCellWidget(row, 4, btn_del)

            self.items.append(
                {
                    "material_id": mid,
                    "category": category,
                    "name": name,
                    "model": model_text,
                    "model_id": int(mid_model) if mid_model is not None else None,
                    "model_label": m_label,
                    "quantity": qty_f,
                }
            )
    
    def select_image(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择图片",
            "",
            "Image Files (*.png *.jpg *.jpeg *.bmp)"
        )

        if file_path:
             self.image_path = file_path
             self.img_label.setText(os.path.basename(file_path))

    def load_warehouses(self):
        """加载仓库到下拉框"""
        warehouses = self.warehouse_mgr.list_warehouses()
        self.from_combo.clear()
        self.to_combo.clear()

        for w in warehouses:
            self.from_combo.addItem(w['name'], w['id'])
            self.to_combo.addItem(w['name'], w['id'])

    def load_materials(self):
        """按类别加载物料到下拉框"""
        self._materials = self.material_mgr.list_materials()
        selected_category = self.category_combo.currentText().strip()
        self.material_combo.clear()
        for m in self._materials:
            category = (m.get("category") or "未分类").strip()
            if selected_category and category != selected_category:
                continue
            display = f"{m['name']} ({m['unit']})"
            self.material_combo.addItem(display, m["id"])
        self.load_models_for_current_material()

    def load_categories(self):
        self.category_combo.clear()
        categories = self.material_mgr.list_categories()
        for c in categories:
            self.category_combo.addItem(c["name"], c["id"])

    def load_models_for_current_material(self):
        material_id = self.material_combo.currentData()
        self.model_combo.clear()
        if material_id is None:
            return
        m = self.material_mgr.get_material(int(material_id))
        if not m:
            return
        model_text = (m.get("model") or "").strip()
        models = [s.strip() for s in model_text.split(",") if s.strip()]
        if models:
            self.model_combo.addItems(models)
    
    def load_warehouse_materials(self, warehouse_id):
        """根据仓库已有库存填充物料下拉（每项携带 material_id 与默认型号）"""
        items = self.warehouse_mgr.list_warehouse_items(warehouse_id)

        self.material_combo.clear()
        self.model_combo.clear()

        for item in items:
            mid = item["material_id"]
            mod = (item.get("model") or "").strip()
            label = item["name"]
            disp = f"{label} ({mod})" if mod else label
            self.material_combo.addItem(disp, (mid, mod))

    def add_material_item(self):
        category = self.category_combo.currentText().strip()
        name = self.material_combo.currentText().strip()
        model_text = self.model_combo.currentText().strip()
        quantity = self.quantity_spin.value()

        if not category:
            QMessageBox.warning(self, "提示", "请先选择类别")
            return

        if not name or not quantity:
            QMessageBox.warning(self, "提示", "请填写完整物料信息")
            return

        data = self.material_combo.currentData()
        material_id = None
        preset_model = ""
        if isinstance(data, tuple) and len(data) >= 2:
            material_id, preset_model = int(data[0]), (data[1] or "").strip()
        elif data is not None:
            try:
                material_id = int(data)
            except (TypeError, ValueError):
                material_id = None

        if material_id is None:
            hit = self.material_mgr.get_material_by_name(name)
            if hit:
                material_id = int(hit["id"])

        if material_id is None:
            QMessageBox.warning(self, "提示", "未找到该物料，请从下拉列表选择或先在物料管理中创建")
            return

        if not model_text and preset_model:
            model_text = preset_model

        model_id = self.material_mgr.find_model_id(material_id, model_text) if model_text else None
        model_label = None if model_id else (model_text or None)

        row = self.item_table.rowCount()
        self.item_table.insertRow(row)

        self.item_table.setItem(row, 0, QTableWidgetItem(category))
        self.item_table.setItem(row, 1, QTableWidgetItem(name))
        self.item_table.setItem(row, 2, QTableWidgetItem(model_text))
        self.item_table.setItem(row, 3, QTableWidgetItem(str(quantity)))

        btn_del = QPushButton("删除")
        btn_del.clicked.connect(lambda _, r=row: self.delete_item_row(r))
        self.item_table.setCellWidget(row, 4, btn_del)

        self.items.append({
            "material_id": material_id,
            "category": category,
            "name": name,
            "model": model_text,
            "model_id": model_id,
            "model_label": model_label,
            "quantity": float(quantity),
        })

    def delete_item_row(self, row):
        """删除某行物料"""
        self.item_table.removeRow(row)
        if row < len(self.items):
            del self.items[row]

    def save_dispatch(self):
        """保存调度记录"""
        from_id = self.from_combo.currentData()
        to_id = self.to_combo.currentData()

        if from_id is None or to_id is None:
            QMessageBox.warning(self, "错误", "请先添加仓库并选择调出/调入仓库")
            return

        if from_id == to_id:
            QMessageBox.warning(self, "错误", "调出仓库和调入仓库不能相同！")
            return

        if not self.items:
            QMessageBox.warning(self, "错误", "请至少添加一种物料！")
            return

        try:
            stored_image = store_dispatch_photo(self.image_path)
        except OSError as e:
            QMessageBox.critical(
                self,
                "图片保存失败",
                f"无法将图片复制到「Picture record」目录：\n{e}",
            )
            return

        payload_items = [
            {
                "material_id": item["material_id"],
                "quantity": float(item["quantity"]),
                "model_id": item.get("model_id"),
                "model_label": item.get("model_label"),
            }
            for item in self.items
        ]

        try:
            if self.edit_record_id:
                self.dispatch_mgr.update_dispatch(
                    self.edit_record_id,
                    from_warehouse_id=from_id,
                    to_warehouse_id=to_id,
                    items=payload_items,
                    image_path=stored_image,
                    executor=self.executor_edit.text().strip(),
                    remarks=self.remarks_edit.text().strip() or None,
                )
                QMessageBox.information(
                    self, "成功", f"调度记录 #{self.edit_record_id} 已更新"
                )
            else:
                record_id = self.dispatch_mgr.add_dispatch(
                    from_warehouse_id=from_id,
                    to_warehouse_id=to_id,
                    items=payload_items,
                    image_path=stored_image,
                    executor=self.executor_edit.text().strip(),
                    remarks=self.remarks_edit.text().strip() or None,
                )
                QMessageBox.information(
                    self, "成功", f"调度记录创建成功！\n记录ID: {record_id}"
                )
            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))

class AddStockDialog(QDialog):
    def __init__(self, parent, warehouse_mgr, material_mgr):
        super().__init__(parent)
        self.warehouse_mgr = warehouse_mgr
        self.material_mgr = material_mgr
        self.setWindowTitle("物料入库")
        self.resize(700, 320)

        # 初始化控件
        self.warehouse_combo = QComboBox()
        self.category_combo = QComboBox()
        self.material_combo = QComboBox()
        self.material_combo.setEditable(True)
        self.model_combo = QComboBox()
        self.model_combo.setEditable(True)
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 999999)
        self.executor_edit = QLineEdit()

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()

        # 添加到表单
        form.addRow("仓库：", self.warehouse_combo)
        form.addRow("类别：", self.category_combo)
        form.addRow("物品：", self.material_combo)
        form.addRow("型号：", self.model_combo)
        form.addRow("数量：", self.quantity_spin)
        form.addRow("执行人：", self.executor_edit)
        layout.addLayout(form)

        # 底部按钮
        btn_layout = QHBoxLayout()
        btn_save = QPushButton("✅ 保存")
        btn_cancel = QPushButton("取消")
        btn_save.clicked.connect(self.save_stock)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

        # 加载数据
        self.load_warehouses()
        self.load_categories()
        self.load_materials()

        # 当类别改变时刷新物品下拉
        self.category_combo.currentTextChanged.connect(self.load_materials)
        self.material_combo.currentIndexChanged.connect(self.load_models)

    def load_warehouses(self):
        self.warehouse_combo.clear()
        for w in self.warehouse_mgr.list_warehouses():
            self.warehouse_combo.addItem(w['name'], w['id'])

    def load_categories(self):
        labels = {(m.get("category") or "未分类") for m in self.material_mgr.list_materials()}
        for c in self.material_mgr.list_categories():
            labels.add(c["name"])
        self.category_combo.clear()
        self.category_combo.addItems(sorted(labels) or ["未分类"])

    def load_materials(self):
        selected_category = self.category_combo.currentText()
        materials = self.material_mgr.list_materials()
        filtered = [
            m for m in materials if (m.get("category") or "未分类") == selected_category
        ]

        self.material_combo.clear()
        for m in filtered:
            self.material_combo.addItem(m["name"], m["id"])

        self.load_models()

    def load_models(self):
        self.model_combo.clear()
        material_id = self.material_combo.currentData()
        if material_id is None:
            return
        m = self.material_mgr.get_material(int(material_id))
        if not m:
            return
        model_text = (m.get("model") or "").strip()
        models = [s.strip() for s in model_text.split(",") if s.strip()]
        if models:
            self.model_combo.addItems(models)

    def save_stock(self):
        warehouse_id = self.warehouse_combo.currentData()
        material_id = self.material_combo.currentData()
        model = (self.model_combo.currentText() or "").strip()
        quantity = float(self.quantity_spin.value())

        if warehouse_id is None or material_id is None:
            QMessageBox.warning(self, "提示", "请选择仓库和物品")
            return

        try:
            inv = InventoryManager(self.material_mgr.db)
            inv.add_stock(int(warehouse_id), int(material_id), model, quantity)
        except Exception as e:
            QMessageBox.critical(self, "入库失败", str(e))
            return

        QMessageBox.information(self, "成功", "入库已保存")
        self.accept()

class DispatchDetailDialog(QDialog):
    def __init__(self, data, parent=None):
        super().__init__(parent)
        self.data = data
        self.material_mgr = parent.material_mgr
        self.dispatch_mgr = parent.dispatch_mgr
        self.setWindowTitle(f"调度详情 - 单号 {data['id']}")
        self.resize(720, 680)
        self.setMinimumSize(560, 480)

        # 设置窗口背景色
        self.setStyleSheet("background-color: #ffffff;")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 14)
        layout.setSpacing(6)

        # --- 标题栏 ---
        title_label = QLabel("🚚 调度记录详情")
        title_label.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 0px;
        """)
        layout.addWidget(title_label)

        # --- 信息展示区（两列排布，减少纵向占位）---
        info_card = QLabel()
        info_card.setTextFormat(Qt.RichText)

        hid = html.escape(str(data.get("id", "")))
        hfrom = html.escape(str(data.get("from", "")))
        hto = html.escape(str(data.get("to", "")))
        hexec = html.escape(str(data.get("executor", "")))
        htime = html.escape(str(data.get("time", "")))

        lab = "color:#909399;font-size:12px;padding:2px 6px 2px 0;white-space:nowrap;"
        val = "color:#303133;font-size:13px;padding:2px 12px 2px 0;"
        content = f"""
        <div style="font-family:'Microsoft YaHei','Segoe UI',sans-serif;">
            <table border="0" cellspacing="0" cellpadding="0" style="line-height:1.25;border-collapse:collapse;">
                <tr>
                    <td style="{lab}">调度编号</td>
                    <td style="{val}font-weight:bold;">{hid}</td>
                    <td style="{lab}">调出仓库</td>
                    <td style="{val}color:#409eff;font-weight:bold;">{hfrom}</td>
                </tr>
                <tr>
                    <td style="{lab}">执行人员</td>
                    <td style="{val}">{hexec}</td>
                    <td style="{lab}">调入仓库</td>
                    <td style="{val}color:#67c23a;font-weight:bold;">{hto}</td>
                </tr>
                <tr>
                    <td style="{lab}">执行时间</td>
                    <td style="{val}" colspan="3">{htime}</td>
                </tr>
            </table>
            <div style="margin:4px 0 6px 0;border-top:1px solid #ebeef5;"></div>
        </div>
        """
        info_card.setText(content)
        layout.addWidget(info_card)

        # --- 物资明细区域 ---
        detail_title = QLabel("物资明细：")
        detail_title.setStyleSheet("""
            font-size: 13px;
            font-weight: bold;
            color: #2c3e50;
            margin-top: 4px;
            padding-top: 0px;
            padding-bottom: 0px;
        """)
        layout.addWidget(detail_title)
        layout.addSpacing(4)

        self.detail_text = QTextEdit()

        def format_materials_from_db() -> str:
            try:
                rid = int(data["id"])
            except (TypeError, ValueError):
                return ""
            items = self.dispatch_mgr.get_dispatch_items(rid)
            if not items:
                return ""
            lines = []
            for it in items:
                name = it.get("material_name") or ""
                mod = (it.get("model_name") or "").strip()
                label = f"{name}({mod})" if mod else name
                qty = it.get("quantity")
                try:
                    qf = float(qty)
                    qty_str = str(int(qf)) if qf == int(qf) else str(qf).rstrip("0").rstrip(".")
                except (TypeError, ValueError):
                    qty_str = str(qty)
                unit = (it.get("unit") or "").strip()
                line = f"• {label} x{qty_str}"
                if unit:
                    line += f" {unit}"
                lines.append(line)
            return "\n".join(lines)

        def format_materials_fallback(text: str) -> str:
            """表格汇总字符串的兜底解析（无 record 或查询失败时使用）"""
            result = []
            for item in re.split(r"[;\n]+", text):
                item = item.strip()
                if not item:
                    continue
                if " x" in item:
                    left, _, right = item.rpartition(" x")
                    name_key = left.split("(")[0].strip()
                    qty = right.replace(".0", "")
                    unit = self._unit_for_material_name(name_key)
                    suf = f" {unit}" if unit else ""
                    result.append(f"• {left.strip()} x{qty}{suf}")
                else:
                    result.append(f"• {item}")
            return "\n".join(result)

        clean_materials = format_materials_from_db()
        if not clean_materials.strip():
            clean_materials = format_materials_fallback(data.get("materials") or "")

        # 每条物料一行一块，带底部分隔线（比纯文本更易区分「每笔」）
        mat_html_parts: List[str] = []
        for raw_line in clean_materials.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            esc = html.escape(line)
            mat_html_parts.append(
                f'<div style="border-bottom:1px solid #ebeef5;padding:8px 4px;font-size:14px;">{esc}</div>'
            )
        if not mat_html_parts:
            mat_html_parts.append(
                '<div style="color:#909399;padding:8px;">（无明细）</div>'
            )
        self.detail_text.setHtml(
            '<div style="font-family:\'Microsoft YaHei\',\'Segoe UI\',sans-serif;">'
            + "".join(mat_html_parts)
            + "</div>"
        )
        self.detail_text.setReadOnly(True)
        
        # 2. 修改列表内容字体：将 font-size 改为 15px 或 16px
        self.detail_text.setStyleSheet("""
            QTextEdit {
                background-color: #f5f7fa;
                border: 1px solid #e4e7ed;
                border-radius: 6px;
                padding: 8px 10px;
                color: #606266;
                font-size: 13px;
                line-height: 1.35;
                font-family: "Segoe UI", "Microsoft YaHei", "Arial Unicode MS", "SimHei";
            }
        """)
        self.detail_text.setMinimumHeight(220)
        layout.addWidget(self.detail_text, stretch=1)

        # --- 底部按钮 ---
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.close_btn = QPushButton("关闭")
        self.close_btn.setCursor(Qt.PointingHandCursor)
        self.close_btn.setStyleSheet("""
            QPushButton {
                background-color: #409eff;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 40px;
                font-size: 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #66b1ff;
            }
        """)
        btn_layout.addWidget(self.close_btn)
        btn_layout.addStretch()

        layout.addLayout(btn_layout)

        self.close_btn.clicked.connect(self.accept)
        
    def _unit_for_material_name(self, name: str) -> str:
        if not name:
            return ""
        m = self.material_mgr.get_material_by_name(name)
        if m:
            return (m.get("unit") or "").strip()
        return ""